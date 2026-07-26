# -*- coding: utf-8 -*-
"""Agent kalibracija MASTER Ocjene sheetova za gin / tequila / brandy.

Postavlja qualityScore, body/sweetness (gin/tequila), i HR komentare na temelju
brand/stil pravila — zamjenjuje rucni Excel korak.

Pokretanje:
  python scripts/calibrate-master.py --category gin
  python scripts/calibrate-master.py --category tequila
  python scripts/calibrate-master.py --category brandy
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent.parent

XLSX = {
    "gin": ROOT / "Gin_Kolekcija_Checklist.xlsx",
    "tequila": ROOT / "Tequila_Kolekcija_Checklist.xlsx",
    "brandy": ROOT / "Konjak_Brandy_Checklist.xlsx",
}

# Brand → (quality nudge absolute or None to keep, note_hr)
GIN_CAL: list[tuple[str, float, str, int, int]] = [
    (r"monkey\s*47", 9.2, "Referentni contemporary gin — gusta botanika; čist ili martini uz Connecticut/blagi Habano.", 3, 2),
    (r"hendrick", 8.8, "Cvjetno-krastavac profil; elegantan neat/martini uz blagu do srednju cigaru.", 3, 2),
    (r"the botanist|botanist", 8.6, "Islay botanika, složen i suhi; martini ili čist uz Connecticut.", 3, 2),
    (r"gin mare", 8.5, "Mediteranska botanika (maslina, timijan); G&T ili neat uz blagi Habano.", 3, 2),
    (r"roku", 8.4, "Japanski contemporary — sakura/yuzu nijanse; čist uz Connecticut.", 3, 2),
    (r"ki\s*no\s*bi|kinobi", 8.3, "Kyoto botanika; fin contemporary uz blagu cigaru.", 3, 2),
    (r"tanqueray\s+no|no\.?\s*ten", 8.4, "Premium London Dry — citrus i borovica; martini klasik.", 3, 2),
    (r"sipsmith", 8.3, "Klasični London Dry craft; martini ili G&T uz Connecticut.", 3, 2),
    (r"plymouth", 8.2, "Mekši od London Dryja; neat/martini uz blagu cigaru.", 3, 2),
    (r"malfy", 7.8, "Talijanski citrusni gin; G&T uz Connecticut.", 3, 2),
    (r"nordes|nordés", 7.9, "Galički mediteranski stil; biljno-citrusni most prema blagoj cigari.", 3, 2),
    (r"citadelle", 7.7, "Francuski dry — uravnotežen uz Connecticut.", 3, 2),
    (r"elephant", 7.6, "Afričke botanike; contemporary uz blagi Habano.", 3, 2),
    (r"beefeater\s*24", 7.8, "Beefeater 24 — japanski sencha nijansa; solidan martini.", 3, 2),
    (r"beefeater", 7.2, "Radni London Dry; G&T svakodnevni pairing s Connecticutom.", 3, 2),
    (r"tanqueray", 7.5, "Klasični London Dry; martini/G&T uz blagu cigaru.", 3, 2),
    (r"bombay", 7.3, "Bombay Sapphire — pristupačan dry uz Connecticut.", 3, 2),
    (r"aviation", 7.6, "Američki contemporary; čist uz blagu Habano.", 3, 2),
    (r"bathtub", 7.4, "Pot-still character; neat oprezno uz srednju cigaru.", 3, 2),
    (r"copperhead", 8.0, "Belgijski premium dry; martini uz Connecticut.", 3, 2),
    (r"silent pool", 8.1, "Engleski botanical; neat/martini uz blagu cigaru.", 3, 2),
    (r"g'?vine", 7.5, "Francuski grape-based gin; cvjetno uz Connecticut.", 3, 2),
    (r"aura|istra|dalma|maraska|badel", 7.4, "HR craft gin — lokalni terroir uz Connecticut/blagi Habano.", 3, 2),
]

TEQUILA_CAL: list[tuple[str, float, str, int, int]] = [
    (r"don julio\s*1942|1942", 9.2, "Extra añejo referent — gusto, desertno; maduro/oscuro, snifter.", 5, 4),
    (r"don julio\s*blanco", 8.5, "Referentni sipping blanco — čista agava i citrus uz laganu do srednju cigaru.", 2, 2),
    (r"don julio\s*reposado", 8.4, "Agava + hrast; uravnotežen uz Habano srednje snage.", 3, 3),
    (r"don julio\s*a[nñ]ejo|don julio\s*anejo", 8.6, "Vanilija i hrast; maduro ga dobro nosi.", 4, 3),
    (r"don julio", 8.3, "Don Julio linija — pouzdan sipping profil uz srednju cigaru.", 3, 3),
    (r"patr[oó]n\s*silver|patron\s*silver", 8.0, "Glatki premium blanco — ulaz za Connecticut/blagi Habano.", 2, 2),
    (r"patr[oó]n\s*reposado|patron\s*reposado", 8.1, "Reposado s blagim hrastom; Habano srednje.", 3, 3),
    (r"patr[oó]n\s*a[nñ]ejo|patron\s*anejo", 8.3, "Añejo — karamela i hrast uz maduro.", 4, 3),
    (r"patr[oó]n|patron", 8.0, "Patrón — premium sipping uz srednju cigaru.", 3, 2),
    (r"herradura\s*blanco", 8.2, "Zemljanija agava; dobro uz srednju cigaru.", 3, 2),
    (r"herradura\s*reposado", 8.3, "Klasični reposado; Habano most.", 3, 3),
    (r"herradura\s*a[nñ]ejo|herradura\s*anejo", 8.4, "Añejo snaga uz maduro.", 4, 3),
    (r"herradura", 8.1, "Herradura — tradicionalni Jalisco profil.", 3, 3),
    (r"espol[oó]n\s*blanco|espolon\s*blanco", 7.6, "Value blanco 100% agave; Connecticut.", 2, 2),
    (r"espol[oó]n\s*reposado|espolon\s*reposado", 7.7, "Value reposado; blagi Habano.", 3, 3),
    (r"espol[oó]n|espolon", 7.5, "Espolòn — pristupačan sipping.", 3, 2),
    (r"casamigos\s*blanco", 7.8, "Mekši blanco; Connecticut.", 2, 2),
    (r"casamigos\s*reposado", 7.9, "Reposado s vanilijom; Habano.", 3, 3),
    (r"casamigos\s*a[nñ]ejo|casamigos\s*anejo", 8.0, "Añejo — desertniji; maduro oprezno.", 4, 3),
    (r"casamigos", 7.8, "Casamigos — gladak stil uz srednju cigaru.", 3, 3),
    (r"clase azul", 8.7, "Premium reposado/XA tier — snifter uz maduro.", 4, 3),
    (r"fortaleza", 8.8, "Craft referent — tahona; blanco/reposado uz Connecticut/Habano.", 3, 2),
    (r"olmeca altos", 7.4, "Solidan value 100% agave; Connecticut.", 2, 2),
    (r"c[oó]digo|codigo", 8.2, "Código — moderni sipping uz Habano.", 3, 3),
    (r"del maguey|montelobos|mezcal", 7.5, "Mezcal sipping — dim uz corojo/oscura; neat.", 4, 2),
]

COGNAC_CAL: list[tuple[str, float, str]] = [
    (r"louis xiii|louis\s*xiii", 9.5, "Ikona XO+ — samo uz najjače maduro/oscuro, ritualni sip."),
    (r"hennessy\s+paradis|paradis", 9.3, "Paradis — ekstremni XO; gusta večer uz oscuro."),
    (r"delamain", 8.8, "Delamain — suhi, elegantni XO; Habano/oscuro."),
    (r"hennessy\s+xo", 8.5, "Hennessy XO — referentni XO uz maduro."),
    (r"r[eé]my\s+martin\s+xo|remy\s+martin\s+xo|r[eé]my\s+xo", 8.4, "Rémy XO — voće i hrast uz maduro."),
    (r"martell\s+xo|martell\s+cordon", 8.3, "Martell XO — suši profil uz Habano/maduro."),
    (r"camus\s+xo|hine\s+xo|frapin|gautier\s+xo", 8.2, "Kućni XO — sipping uz srednje-jake cigare."),
    (r"hennessy\s+vsop|r[eé]my.*vsop|martell\s+vsop|1738", 7.8, "VSOP — cisto ili kocka leda uz Connecticut/Habano."),
    (r"hennessy\s+vs\b|r[eé]my.*\bvs\b|martell\s+vs\b", 6.8, "VS — ulazni konjak; blaga cigara, ne maduro bomba."),
    (r"courvoisier", 7.6, "Courvoisier — uravnotežen cognac uz srednju cigaru."),
]


def _match_cal(name: str, rules: list) -> tuple | None:
    low = name.lower()
    for rule in rules:
        if re.search(rule[0], low, re.I):
            return rule
    return None


def calibrate_gin(ws) -> int:
    n = 0
    for row in ws.iter_rows(min_row=3):
        name = row[0].value
        quality = row[1].value
        if not name or quality is None:
            continue
        name_str = str(name)
        if name_str.startswith(("VRH", "Odlican", "Value")):
            continue
        hit = _match_cal(name_str, GIN_CAL)
        if not hit:
            # Soft floor for unlabeled pairable gins
            if float(quality) < 6.5:
                row[1].value = 6.5
                n += 1
            if not row[10].value or str(row[10].value).startswith("Heuristika"):
                style = str(row[3].value or "contemporary")
                row[10].value = f"Sipping gin ({style}) — neat/martini uz Connecticut ili blagi Habano."
                n += 1
            continue
        _, q, note, body, sweet = hit
        row[1].value = q
        row[5].value = body
        row[6].value = sweet
        row[10].value = note
        n += 1
    return n


def calibrate_tequila(ws) -> int:
    n = 0
    for row in ws.iter_rows(min_row=3):
        name = row[0].value
        quality = row[1].value
        if not name or quality is None:
            continue
        name_str = str(name)
        if name_str.startswith(("VRH", "Odlican", "Value")):
            continue
        hit = _match_cal(name_str, TEQUILA_CAL)
        if not hit:
            if float(quality) < 6.5:
                row[1].value = 6.5
                n += 1
            if not row[10].value or str(row[10].value).startswith("Heuristika"):
                tier = str(row[4].value or "tequila")
                row[10].value = f"Sipping tequila ({tier}) — čisto uz cigaru po snazi odležavanja."
                n += 1
            continue
        _, q, note, body, sweet = hit
        row[1].value = q
        row[5].value = body
        row[6].value = sweet
        row[10].value = note
        n += 1
    return n


def calibrate_brandy(ws) -> int:
    """Only bump cognac-class rows that still have heuristic notes or weak scores."""
    n = 0
    for row in ws.iter_rows(min_row=3):
        name = row[0].value
        quality = row[1].value
        if not name or quality is None:
            continue
        name_str = str(name)
        if name_str.startswith(("VRH", "Odlican", "Value")):
            continue
        cat = str(row[3].value or "").lower()
        hit = _match_cal(name_str, COGNAC_CAL)
        if not hit:
            continue
        # Focus on cognac / known cognac brands
        if cat and "cognac" not in cat and "konjak" not in cat:
            if not re.search(r"hennessy|martell|r[eé]my|camus|delamain|hine|courvoisier", name_str, re.I):
                continue
        _, q, note = hit
        old_q = float(quality)
        # Raise toward calibrated if heuristic underestimates; never lower curated high scores much
        if old_q < q:
            row[1].value = q
            n += 1
        comment = row[9].value if len(row) > 9 else None
        if not comment or str(comment).startswith("Heuristika"):
            row[9].value = note
            n += 1
    return n


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--category", required=True, choices=["gin", "tequila", "brandy"])
    args = ap.parse_args()

    path = XLSX[args.category]
    if not path.exists():
        print(f"Missing {path}")
        return 1
    wb = openpyxl.load_workbook(path)
    if "MASTER Ocjene" not in wb.sheetnames:
        print("No MASTER Ocjene sheet")
        return 1
    ws = wb["MASTER Ocjene"]
    if args.category == "gin":
        n = calibrate_gin(ws)
    elif args.category == "tequila":
        n = calibrate_tequila(ws)
    else:
        n = calibrate_brandy(ws)
    wb.save(path)
    print(f"calibrated {args.category}: {n} cell updates -> {path.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
