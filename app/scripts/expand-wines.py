# -*- coding: utf-8 -*-
"""Oporavi/proširi wines.json iz Vino_Kolekcija_Checklist.xlsx (MASTER/Sva vina)
ili iz app/scripts/data/wine_catalog_seed.json ako postoji.

Preferirani put: seed JSON (puni Drink zapisi). Fallback: Excel + stari wines.json.

Pokretanje (iz root ili app/):
  python app/scripts/expand-wines.py
"""
from __future__ import annotations

import json
import re
import sys
from copy import deepcopy
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
APP = ROOT / "app"
WINES = APP / "src" / "data" / "wines.json"
SEED = APP / "scripts" / "data" / "wine_catalog_seed.json"
CHECKLIST = ROOT / "Vino_Kolekcija_Checklist.xlsx"

CLEAN = {
    "hr": "Standardno vinarstvo (sulfiti)",
    "en": "Standard winemaking (sulphites)",
}
FORTIFIED = {
    "hr": "Fortificirano vinskim destilatom; slatkoca iz zaustavljene fermentacije",
    "en": "Fortified with grape spirit; sweetness from arrested fermentation",
}

STYLE_FROM_LABEL = {
    "porto tawny": "port-tawny",
    "porto ruby": "port-ruby",
    "port-tawny": "port-tawny",
    "port-ruby": "port-ruby",
    "sherry suhi": "sherry-dry",
    "sherry slatki": "sherry-sweet",
    "sherry-dry": "sherry-dry",
    "sherry-sweet": "sherry-sweet",
    "madeira": "madeira",
    "prošek": "prosek",
    "prosek": "prosek",
    "desertno": "dessert-wine",
    "dessert-wine": "dessert-wine",
    "crveno puno": "red-full",
    "red-full": "red-full",
    "crveno srednje": "red-medium",
    "red-medium": "red-medium",
    "bijelo svježe": "white-fresh",
    "bijelo bogato": "white-rich",
    "white-fresh": "white-fresh",
    "white-rich": "white-rich",
    "pjenušavo": "sparkling",
    "sparkling": "sparkling",
}


def slugify(name: str) -> str:
    s = name.lower()
    s = s.replace("&", " and ")
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return "wine-" + s.strip("-")[:60]


def parse_price(cell) -> dict | None:
    if cell is None:
        return None
    t = str(cell).replace("€", "").replace(",", ".").strip()
    m = re.findall(r"\d+(?:\.\d+)?", t)
    if not m:
        return None
    nums = [float(x) for x in m]
    return {"min": nums[0], "max": nums[-1]}


def parse_style_region(cell: str) -> tuple[str, str]:
    raw = (cell or "").strip()
    if " / " in raw:
        style_part, region = raw.split(" / ", 1)
    else:
        style_part, region = raw, ""
    key = style_part.strip().lower()
    style = STYLE_FROM_LABEL.get(key, key.replace(" ", "-"))
    return style, region.strip()


def additive_for(style: str) -> tuple[str, dict, str]:
    if style in ("port-tawny", "port-ruby", "sherry-dry", "sherry-sweet", "madeira", "prosek"):
        return "fortified", FORTIFIED, "Standard kategorije porto/sherry/madeira"
    return "clean", CLEAN, "Standard kategorije vino"


def serve_for(style: str) -> str:
    return {
        "port-tawny": "Casa za porto, 14-16 C",
        "port-ruby": "Casa za porto, 14-16 C",
        "sherry-dry": "Casa za sherry, 12-14 C",
        "sherry-sweet": "Casa za sherry, 14-16 C",
        "madeira": "Casa za madeira, 14-16 C",
        "prosek": "Mala casa, 12-14 C",
        "dessert-wine": "Ohladjeno, 8-10 C",
        "red-full": "Velika casa, 16-18 C",
        "red-medium": "Casa, 15-17 C",
        "white-fresh": "Dobro ohladjeno, 6-8 C",
        "white-rich": "Ohladjeno, 8-10 C",
        "sparkling": "Dobro ohladjeno, 6-8 C",
    }.get(style, "Casa, 14-16 C")


def from_seed() -> list[dict]:
    return json.loads(SEED.read_text(encoding="utf-8"))


def from_excel(old: list[dict]) -> list[dict]:
    from openpyxl import load_workbook

    by_name = {w["name"]: w for w in old}
    wb = load_workbook(CHECKLIST, read_only=True, data_only=True)

    # MASTER: Vino, Kval, Stil/Regija, Tijelo, Slatkoca, Note(tags), Cijena, Ducan
    master: dict[str, dict] = {}
    if "MASTER Ocjene" in wb.sheetnames:
        for row in wb["MASTER Ocjene"].iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            if not name or row[1] is None:
                continue
            if any(name.startswith(p) for p in ("VRH", "Odli", "Value", "Ostalo")):
                continue
            style, region = parse_style_region(str(row[2] or ""))
            tags = [t.strip() for t in str(row[5] or "").split(",") if t.strip()]
            master[name] = {
                "qs": row[1],
                "style": style,
                "region": region,
                "body": int(row[3] or 3),
                "sweetness": int(row[4] or 1),
                "tags": tags,
                "price": parse_price(row[6]),
                "shop": str(row[7] or "Vinoteke"),
            }

    # Sva vina: prose notes + ABV
    ranked: dict[str, dict] = {}
    sheet = "Sva vina (rang)" if "Sva vina (rang)" in wb.sheetnames else None
    if sheet:
        for row in wb[sheet].iter_rows(min_row=3, values_only=True):
            if not row or not row[1]:
                continue
            name = str(row[1]).strip()
            ranked[name] = {
                "notes_hr": str(row[4] or ""),
                "abv": row[7],
                "price": parse_price(row[5]),
                "shop": str(row[6] or "Vinoteke"),
                "qs": row[2],
                "style_reg": str(row[3] or ""),
            }

    names = list(master.keys()) if master else list(ranked.keys())
    out: list[dict] = []
    seen: set[str] = set()
    for name in names:
        m = master.get(name, {})
        r = ranked.get(name, {})
        style = m.get("style")
        region = m.get("region", "")
        if not style and r.get("style_reg"):
            style, region = parse_style_region(r["style_reg"])
        style = style or "red-full"
        tags = m.get("tags") or (by_name.get(name) or {}).get("flavorTags") or ["voce"]
        # drop prose-like tags
        tags = [t for t in tags if len(t) <= 24 and " " not in t or t in (
            "suho-voce", "tamno-voce", "tropsko-voce", "trava-slatka", "zacini-slatki", "ester-funk"
        )]
        if not tags:
            tags = ["voce"]
        old_hit = by_name.get(name)
        wid = (old_hit or {}).get("id") or slugify(name)
        base = 2
        while wid in seen:
            wid = f"{slugify(name)}-{base}"
            base += 1
        seen.add(wid)
        additive, detail, source = additive_for(style)
        abv = r.get("abv")
        if abv is None:
            abv = (old_hit or {}).get("abv") or 13.0
        notes_hr = r.get("notes_hr") or (old_hit or {}).get("notes", {}).get("hr") or ", ".join(tags)
        notes_en = (old_hit or {}).get("notes", {}).get("en") or notes_hr
        qs = m.get("qs", r.get("qs", 7.0))
        out.append({
            "id": wid,
            "category": "wine",
            "name": name,
            "style": style,
            "region": region or (old_hit or {}).get("region") or "",
            "abv": float(abv),
            "body": int(m.get("body") or (old_hit or {}).get("body") or 3),
            "sweetness": int(m.get("sweetness") or (old_hit or {}).get("sweetness") or 1),
            "flavorTags": tags[:6],
            "additiveStatus": additive,
            "additiveDetail": detail,
            "additiveSource": source,
            "qualityScore": float(qs) if qs is not None else 7.0,
            "priceEUR": m.get("price") or r.get("price") or {"min": 15.0, "max": 30.0},
            "priceApprox": True,
            "shopHR": m.get("shop") or r.get("shop") or "Vinoteke",
            "status": None,
            "pairable": True,
            "serving": {"best": serve_for(style)},
            "cigarHint": None,
            "priceUrl": (old_hit or {}).get("priceUrl"),
            "notes": {
                "hr": str(notes_hr).replace("\u2014", ",").replace("—", ","),
                "en": str(notes_en).replace("\u2014", ",").replace("—", ","),
            },
        })
    wb.close()
    return out


def main() -> None:
    old = json.loads(WINES.read_text(encoding="utf-8"))
    if SEED.exists():
        new = from_seed()
        print(f"from seed: {len(new)}")
    elif CHECKLIST.exists():
        new = from_excel(old)
        print(f"from excel: {len(new)}")
    else:
        print("No seed or checklist", file=sys.stderr)
        sys.exit(1)
    ids = [w["id"] for w in new]
    if len(ids) != len(set(ids)):
        print("WARNING: duplicate ids", file=sys.stderr)
    WINES.write_text(json.dumps(new, ensure_ascii=False, indent=1) + "\n", encoding="utf-8")
    print(f"Wrote {WINES} count={len(new)}")


if __name__ == "__main__":
    main()
