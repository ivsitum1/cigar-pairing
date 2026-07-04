# -*- coding: utf-8 -*-
"""Ekstrakcija Whisky_Kolekcija_Checklist.xlsx -> whiskies.json za pairing aplikaciju.

Cita MASTER Ocjene, Serviranje + Cigare i Katalog allez+ecuga sheetove.

Pokretanje: python scripts/excel-to-whisky-json.py
"""
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

from whisky_shared import (
    additive_status,
    catalog_index,
    cigar_hint_for_style,
    detect_expression_type,
    detect_style_region,
    extract_abv,
    find_best_catalog_match,
    is_pairable,
    match_tokens,
    normalize_region,
    parse_price_eur,
    serving_for_style,
    slugify,
)

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT.parent / "Whisky_Kolekcija_Checklist.xlsx"
OUT = ROOT / "src" / "data" / "whiskies.json"

SERVE_SCORE = {"++": 3, "+": 2, "~": 1, "x": 0}


def build_catalog_index(wb) -> list[dict]:
    index = []
    for row in wb["Katalog allez+ecuga"].iter_rows(min_row=3, values_only=True):
        name, price, shop, url = row[0], row[1], row[2], row[3]
        if not name or not url:
            continue
        index.append({
            "name": str(name),
            "url": str(url),
            "shop": str(shop or ""),
            "price_eur": parse_price_eur(price),
            "tokens": match_tokens(str(name)),
        })
    return index


def build_serve_index(wb) -> list[dict]:
    rows = []
    ws = wb["Serviranje + Cigare"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        name, neat, water, rocks, best, hint = row[0], row[1], row[2], row[3], row[4], row[5]
        if not name or not best:
            continue
        rows.append({
            "tokens": match_tokens(str(name)),
            "serving": {
                "neat": SERVE_SCORE.get(str(neat).strip(), None) if neat else None,
                "water": SERVE_SCORE.get(str(water).strip(), None) if water else None,
                "rocks": SERVE_SCORE.get(str(rocks).strip(), None) if rocks else None,
                "highball": 0,
                "cola": 0,
                "best": str(best),
            },
            "cigarHint": str(hint) if hint else None,
        })
    return rows


def country_from_region(region: str) -> str:
    if "Škotska" in region or "Scotland" in region:
        return "Škotska"
    if "Irska" in region or "Ireland" in region:
        return "Irska"
    if "SAD" in region or "Kentucky" in region or "Tennessee" in region:
        return "SAD"
    if "Japan" in region:
        return "Japan"
    if "Kanada" in region:
        return "Kanada"
    return region.split(",")[-1].strip() if "," in region else region


def extract_whiskies(wb) -> list[dict]:
    catalog = build_catalog_index(wb)
    serve_rows = build_serve_index(wb)
    ws = wb["MASTER Ocjene"]
    items = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        name, quality, region_col = row[0], row[1], row[2]
        if not name or quality is None:
            continue
        if str(name).startswith("VRH") or str(name).startswith("Odličan") or str(name).startswith("Value"):
            continue

        coloring = str(row[3] or "unknown")
        filt = str(row[4] or "unknown")
        expr = str(row[5] or detect_expression_type(str(name)))
        price_raw = row[6]
        shop = str(row[7] or "")
        comment = str(row[9] or "")

        style, region, body, sweetness, tags = detect_style_region(str(name))
        region_display = normalize_region(str(region_col or region), style)

        abv = extract_abv(str(name))
        additive = additive_status(coloring, expr)
        quality_f = float(quality)

        # serving match
        my_tokens = match_tokens(str(name))
        best_match, best_score = None, 0
        for sr in serve_rows:
            overlap = len(my_tokens & sr["tokens"])
            if overlap > best_score:
                best_match, best_score = sr, overlap
        if best_match and best_score >= 2:
            serving = dict(best_match["serving"])
            cigar_hint = best_match["cigarHint"]
        else:
            serving = serving_for_style(style, abv, expr)
            cigar_hint = cigar_hint_for_style(style)

        cat = find_best_catalog_match(str(name), catalog)
        price_eur = parse_price_eur(price_raw)
        price_url = cat["url"] if cat else None
        if price_eur is None and cat and cat.get("price_eur"):
            price_eur = cat["price_eur"]
        if cat and not shop:
            shop = cat.get("shop", "")

        if abv and abv >= 55 and "overproof" not in tags:
            tags = list(tags) + ["overproof"]

        items.append({
            "id": "wh-" + slugify(str(name)),
            "category": "whisky",
            "name": str(name),
            "style": style,
            "region": region_display,
            "country": country_from_region(region_display),
            "abv": abv,
            "body": body,
            "sweetness": sweetness,
            "flavorTags": tags,
            "additiveStatus": additive,
            "additiveDetail": f"{coloring} / {filt}",
            "additiveSource": "Excel MASTER",
            "qualityScore": quality_f,
            "priceEUR": {"min": price_eur, "max": price_eur} if price_eur else None,
            "priceApprox": price_eur is None,
            "shopHR": shop or "allez.hr",
            "status": str(row[8]).strip() if row[8] and str(row[8]).strip() not in ("-", "") else None,
            "pairable": is_pairable(expr, style, quality_f),
            "serving": serving,
            "cigarHint": cigar_hint,
            "priceUrl": price_url,
            "notes": {"hr": comment, "en": ""},
        })
    return items


def main() -> int:
    if not XLSX.exists():
        print(f"Missing {XLSX} — run build-whisky-excel.py first")
        return 1
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    whiskies = extract_whiskies(wb)
    wb.close()
    OUT.write_text(json.dumps(whiskies, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"whiskies.json: {len(whiskies)} whiskyja")
    with_url = sum(1 for w in whiskies if w.get("priceUrl"))
    print(f"  priceUrl: {with_url}/{len(whiskies)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
