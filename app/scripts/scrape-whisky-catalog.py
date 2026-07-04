# -*- coding: utf-8 -*-
"""Scrape whisky catalog from allez.hr + ecuga.com.

Output: app/scripts/output/whisky_catalog_raw.json
        (optional) updates Katalog sheet if Whisky_Kolekcija_Checklist.xlsx exists

Pokretanje: python scripts/scrape-whisky-catalog.py
            python scripts/scrape-whisky-catalog.py --allez-only
"""
from __future__ import annotations

import argparse
import json
import re
import time
import urllib.error
import urllib.request
from pathlib import Path

from whisky_shared import ECUGA_CATEGORIES, format_price_eur

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = Path(__file__).resolve().parent / "output"
OUT_JSON = OUT_DIR / "whisky_catalog_raw.json"
XLSX = ROOT.parent / "Whisky_Kolekcija_Checklist.xlsx"

UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
ALLEZ_LIST = "https://allez.hr/shop/whiskey"


def fetch_html(url: str) -> str:
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=45) as resp:
        return resp.read().decode("utf-8", "replace")


def scrape_allez_page(html: str) -> list[dict]:
    items: list[dict] = []
    parts = html.split('class="product-image"')
    for part in parts[1:]:
        href_m = re.search(r'<a href="([^"]+)"', part)
        alt_m = re.search(r'alt="([^"]+)"', part)
        if not href_m or not alt_m:
            continue
        href = href_m.group(1)
        if not href.startswith("http"):
            href = "https://allez.hr" + href
        name = alt_m.group(1).strip()
        price_m = re.search(r"([\d]{1,4}[.,][\d]{2})\s*€", part[:1500])
        price = None
        if price_m:
            price = float(price_m.group(1).replace(",", "."))
        items.append(
            {
                "name": name,
                "price_eur": price,
                "shop": "allez.hr",
                "url": href,
                "source": "allez",
            }
        )
    return items


def allez_max_page(html: str) -> int:
    pages = [int(p) for p in re.findall(r"whiskey\?page=(\d+)", html)]
    return max(pages) if pages else 1


def scrape_allez() -> list[dict]:
    print("Scraping allez.hr …")
    first = fetch_html(ALLEZ_LIST)
    max_page = allez_max_page(first)
    print(f"  pages: {max_page}")
    all_items: list[dict] = []
    seen_urls: set[str] = set()
    for page in range(1, max_page + 1):
        url = ALLEZ_LIST if page == 1 else f"{ALLEZ_LIST}?page={page}"
        html = first if page == 1 else fetch_html(url)
        batch = scrape_allez_page(html)
        for item in batch:
            if item["url"] in seen_urls:
                continue
            seen_urls.add(item["url"])
            all_items.append(item)
        if page % 10 == 0 or page == max_page:
            print(f"  page {page}/{max_page} -> {len(all_items)} unique")
        time.sleep(0.35)
    print(f"allez.hr: {len(all_items)} stavki")
    return all_items


def scrape_ecuga() -> list[dict]:
    from playwright.sync_api import sync_playwright

    print("Scraping ecuga.com (Playwright + GraphQL intercept) ...")
    all_items: list[dict] = []
    seen_slugs: set[str] = set()
    captured: list[dict] = []

    def on_response(response) -> None:
        if "graphql" not in response.url or response.status != 200:
            return
        try:
            body = response.json()
        except Exception:
            return
        products = (body.get("data") or {}).get("products")
        if products and products.get("edges"):
            captured.append(products)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.on("response", on_response)

        page.goto("https://ecuga.com/katalog/whisky", wait_until="domcontentloaded", timeout=60000)
        for sel in ['button:has-text("Dopusti sve")', 'button:has-text("Dopusti selektirane")']:
            try:
                page.locator(sel).first.click(timeout=4000)
                break
            except Exception:
                pass

        for slug, _cat_id, label in ECUGA_CATEGORIES:
            captured.clear()
            url = f"https://ecuga.com/katalog/whisky/{slug}"
            page.goto(url, wait_until="networkidle", timeout=90000)
            page.wait_for_timeout(2500)
            for _ in range(6):
                page.mouse.wheel(0, 3000)
                page.wait_for_timeout(700)

            cat_count = 0
            for batch in captured:
                for edge in batch.get("edges") or []:
                    node = edge.get("node") or {}
                    slug_p = node.get("slug") or ""
                    if not slug_p or slug_p in seen_slugs:
                        continue
                    seen_slugs.add(slug_p)
                    name = (node.get("name") or "").strip()
                    pricing = ((node.get("pricing") or {}).get("priceRange") or {}).get("start") or {}
                    gross = pricing.get("gross") or {}
                    amount = gross.get("amount")
                    all_items.append(
                        {
                            "name": name,
                            "price_eur": float(amount) if amount is not None else None,
                            "shop": "ecuga.com",
                            "url": f"https://ecuga.com/katalog/whisky/{slug}/{slug_p}",
                            "source": "ecuga",
                            "ecuga_category": label,
                            "ecuga_slug": slug_p,
                        }
                    )
                    cat_count += 1
            print(f"  {label}: {cat_count}")
        browser.close()

    print(f"ecuga.com: {len(all_items)} stavki")
    return all_items


def write_katalog_sheet(entries: list[dict]) -> None:
    if not XLSX.exists():
        return
    import openpyxl

    wb = openpyxl.load_workbook(XLSX)
    if "Katalog allez+ecuga" not in wb.sheetnames:
        wb.close()
        return
    ws = wb["Katalog allez+ecuga"]
    # clear old rows from row 3
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)
    for item in entries:
        ws.append(
            [
                item["name"],
                format_price_eur(item.get("price_eur")),
                item["shop"],
                item["url"],
            ]
        )
    wb.save(XLSX)
    wb.close()
    print(f"Updated Katalog sheet in {XLSX.name}")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--allez-only", action="store_true")
    parser.add_argument("--ecuga-only", action="store_true")
    args = parser.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    entries: list[dict] = []
    if args.ecuga_only and OUT_JSON.exists():
        entries = json.loads(OUT_JSON.read_text(encoding="utf-8"))
        entries = [e for e in entries if e.get("source") != "ecuga"]
    elif not args.ecuga_only:
        entries.extend(scrape_allez())
    if not args.allez_only:
        try:
            entries.extend(scrape_ecuga())
        except Exception as exc:
            print(f"WARNING: ecuga scrape failed: {exc}")

    OUT_JSON.write_text(
        json.dumps(entries, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Wrote {OUT_JSON} ({len(entries)} rows)")
    if XLSX.exists():
        write_katalog_sheet(entries)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
