#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Popravi netočne podatke u Excel checklistama (Serviranje + Cigare, brandy MASTER).

Koristi seed/serve_corrections.json kao izvor istine. Excel datoteke su git-ignorirane
u rootu repozitorija.

  python scripts/export-serve-corrections.py   # osvježi corrections iz app JSON-a
  python scripts/fix-excel-data.py             # primijeni na *.xlsx u rootu
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

from brandy_shared import detect_style_region, normalize_region
from serve_shared import (
    RUM_SERVE_PROFILES,
    find_correction,
    load_corrections,
    normalize_hint,
)

ROOT = Path(__file__).resolve().parent.parent
REPO = ROOT.parent

WORKBOOKS = {
    "rum": REPO / "Rum_Kolekcija_Checklist.xlsx",
    "whisky": REPO / "Whisky_Kolekcija_Checklist.xlsx",
    "brandy": REPO / "Konjak_Brandy_Checklist.xlsx",
}

SERVE_SHEET = "Serviranje + Cigare"
MASTER_SHEET = "MASTER Ocjene"


def fix_serve_sheet(ws, corrections: dict, *, prepend_rum_profiles: bool, short_layout: bool = False) -> int:
    """Vrati broj ispravljenih redova."""
    fixed = 0
    rows = list(ws.iter_rows(min_row=3, values_only=False))
    hint_idx = 5 if short_layout else 7
    if prepend_rum_profiles and rows:
        # Zamijeni profilne redove (prvih len(RUM_SERVE_PROFILES) nakon headera)
        for i, prof in enumerate(RUM_SERVE_PROFILES):
            row_idx = 3 + i
            if row_idx > ws.max_row:
                ws.append([None] * 8)
            for col, key in enumerate(
                ["name", "neat", "water", "rocks", "highball", "cola", "best", "cigarHint"], 1
            ):
                ws.cell(row=row_idx, column=col, value=prof.get(key))
            fixed += 1

    for row in rows:
        name_cell = row[0]
        name = name_cell.value
        if not name or not str(name).strip():
            continue
        name_str = str(name).strip()
        corr = find_correction(name_str, corrections)
        hint_cell = row[hint_idx] if len(row) > hint_idx else None
        old_hint = str(hint_cell.value).strip() if hint_cell and hint_cell.value else ""

        if corr:
            row[1].value = corr.get("neat", row[1].value)
            row[2].value = corr.get("water", row[2].value)
            row[3].value = corr.get("rocks", row[3].value)
            if not short_layout:
                if len(row) > 4:
                    row[4].value = corr.get("highball", row[4].value if len(row) > 4 else "x")
                if len(row) > 5:
                    row[5].value = corr.get("cola", row[5].value if len(row) > 5 else "x")
                if len(row) > 6:
                    row[6].value = corr.get("best", row[6].value)
                if len(row) > 7:
                    row[7].value = corr.get("cigarHint")
            else:
                if len(row) > 4:
                    row[4].value = corr.get("best", row[4].value)
                if len(row) > 5:
                    row[5].value = corr.get("cigarHint")
            fixed += 1
        elif old_hint:
            new_hint = normalize_hint(old_hint)
            if new_hint != old_hint and hint_cell is not None:
                hint_cell.value = new_hint
                fixed += 1

    return fixed


def fix_brandy_master(ws) -> int:
    """Ispravi style/region u MASTER Ocjene prema brandy_shared (XO regex fix)."""
    fixed = 0
    for row in ws.iter_rows(min_row=3):
        name = row[0].value
        if not name or str(name).startswith("VRH"):
            continue
        name_str = str(name)
        style, region, *_ = detect_style_region(name_str)
        region_display = normalize_region(str(row[2].value or region), style)
        old_style = str(row[3].value or "") if len(row) > 3 else ""
        old_region = str(row[2].value or "")
        if old_style != style:
            if len(row) > 3:
                row[3].value = style
            fixed += 1
        if old_region != region_display:
            row[2].value = region_display
            fixed += 1
    return fixed


def process_workbook(kind: str, path: Path, corrections: dict) -> bool:
    if not path.exists():
        print(f"  [skip] {path.name} — nema datoteke")
        return False
    wb = openpyxl.load_workbook(path)
    total = 0
    if SERVE_SHEET in wb.sheetnames:
        n = fix_serve_sheet(
            wb[SERVE_SHEET],
            corrections,
            prepend_rum_profiles=(kind == "rum"),
            short_layout=(kind != "rum"),
        )
        print(f"  {path.name} / {SERVE_SHEET}: {n} redova")
        total += n
    if kind == "brandy" and MASTER_SHEET in wb.sheetnames:
        n = fix_brandy_master(wb[MASTER_SHEET])
        print(f"  {path.name} / {MASTER_SHEET}: {n} brandy polja")
        total += n
    if total:
        wb.save(path)
        print(f"  -> spremljeno {path}")
    else:
        print(f"  -> bez promjena {path}")
    return True


def main() -> None:
    corrections_path = Path(__file__).resolve().parent / "seed" / "serve_corrections.json"
    if not corrections_path.exists():
        print("Nema serve_corrections.json — pokreni export-serve-corrections.py")
        sys.exit(1)
    corrections = load_corrections()
    print(f"Učitano {len(corrections.get('by_name', {}))} ispravki po imenu")
    any_found = False
    for kind, path in WORKBOOKS.items():
        print(f"\n{kind}:")
        if process_workbook(kind, path, corrections):
            any_found = True
    if not any_found:
        print("\nNijedan Excel nije pronađen u rootu repozitorija.")
        print("Kad dodaš Rum_Kolekcija_Checklist.xlsx / Whisky_* / Konjak_*, pokreni ponovo.")
        sys.exit(0)


if __name__ == "__main__":
    main()
