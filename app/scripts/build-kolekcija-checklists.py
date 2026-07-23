# -*- coding: utf-8 -*-
"""Nadopuni/ napravi Kolekcija Checklist Excelle za sva pića + cigare.

Izvor istine za kurirane sheetove: app/src/data/*.json
Zadržava se struktura checklista (MASTER, Po tipu, Serviranje, Kolekcija, Vodič…).

Pokretanje (iz app/):
  python scripts/build-kolekcija-checklists.py
  python scripts/build-kolekcija-checklists.py --only rum|whisky|brandy|wine|gin|coffee|tequila|cigar
"""
from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from serve_shared import resolve_serve_hint
from whisky_shared import (
    additive_status,
    cigar_hint_for_style,
    detect_coloring,
    detect_expression_type,
    detect_filter,
    detect_style_region,
    estimate_quality,
    extract_abv,
    format_price_eur,
    is_pairable,
    match_tokens,
    normalize_region,
    serving_for_style,
    token_overlap,
)

ROOT = Path(__file__).resolve().parent.parent  # app/
DATA = ROOT / "src" / "data"
OUT = ROOT.parent
WHISKY_RAW = Path(__file__).resolve().parent / "output" / "whisky_catalog_raw.json"
BRANDY_RAW = Path(__file__).resolve().parent / "output" / "brandy_catalog_raw.json"

HEADER_FILL = PatternFill("solid", fgColor="2C211A")
HEADER_FONT = Font(bold=True, color="C9A35C")
TITLE_FONT = Font(bold=True, size=12, color="9C4433")
SECTION_FONT = Font(bold=True, color="9C4433")

SERVE_MARK = {3: "++", 2: "+", 1: "~", 0: "x"}


def style_header(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center")


def set_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def price_str(p) -> str:
    if p is None:
        return "provjeriti"
    if isinstance(p, (int, float)):
        return f"{p:.0f} €"
    mn, mx = p.get("min"), p.get("max")
    if mn is None:
        return "provjeriti"
    if mx is None or abs(mn - mx) < 0.01:
        return f"{mn:.2f} €".replace(".", ",")
    return f"{mn:.0f}-{mx:.0f} €"


def notes_hr(d: dict) -> str:
    n = d.get("notes") or {}
    if isinstance(n, dict):
        return n.get("hr") or ""
    return str(n or "")


def additive_display(d: dict) -> str:
    detail = d.get("additiveDetail")
    if isinstance(detail, dict) and detail.get("hr"):
        return str(detail["hr"])
    if isinstance(detail, str) and detail.strip():
        return detail
    return str(d.get("additiveStatus") or "")


def cigar_hint_hr(d: dict) -> str:
    h = d.get("cigarHint")
    if isinstance(h, dict):
        return h.get("hr") or ""
    return str(h or "")


def serve_cell(serving: dict | None, key: str) -> str:
    if not serving:
        return ""
    if key == "best":
        return str(serving.get("best") or "")
    val = serving.get(key)
    if val is None:
        return ""
    try:
        return SERVE_MARK.get(int(val), str(val))
    except (TypeError, ValueError):
        return str(val)


def quality_tier_label(q: float) -> str:
    if q >= 8:
        return "VRH za cigaru (8.0-10)"
    if q >= 7:
        return "Odličan sipper (7-8)"
    if q >= 6:
        return "Value / solidan (6-7)"
    return "Ostalo (<6)"


def replace_sheet(wb: openpyxl.Workbook, title: str, index: int | None = None):
    if title in wb.sheetnames:
        del wb[title]
    if index is None:
        return wb.create_sheet(title)
    return wb.create_sheet(title, index)


def load_json(name: str) -> list | dict:
    return json.loads((DATA / name).read_text(encoding="utf-8"))


# --------------------------------------------------------------------------- rum


def _old_rum_status_map(wb) -> dict[str, str]:
    """Sačuvaj Status markere (META/IMAS/PROBAO) po tokenima."""
    out: dict[str, str] = {}
    if "MASTER Ocjene" not in wb.sheetnames:
        return out
    for row in wb["MASTER Ocjene"].iter_rows(min_row=3, values_only=True):
        name, quality, status = row[0], row[1], row[7] if len(row) > 7 else None
        if not name or quality is None:
            continue
        if status and str(status).strip() not in ("", "-"):
            key = "|".join(sorted(match_tokens(str(name))))
            out[key] = str(status).strip()
    return out


def refresh_rum_checklist() -> None:
    path = OUT / "Rum_Kolekcija_Checklist.xlsx"
    if not path.exists():
        raise SystemExit(f"Missing {path}")
    rums: list[dict] = load_json("rums.json")
    rums = sorted(rums, key=lambda d: (-(d.get("qualityScore") or 0), d.get("name", "")))

    wb = openpyxl.load_workbook(path)
    status_map = _old_rum_status_map(wb)

    # --- MASTER (zadrži stare stupce, dodaj nove na kraju) ---
    idx = wb.sheetnames.index("MASTER Ocjene") if "MASTER Ocjene" in wb.sheetnames else 0
    ws = replace_sheet(wb, "MASTER Ocjene", idx)
    ws.append([
        "MASTER — REKALIBRIRANO za SIPPING UZ CIGARU. Rangirano po kvaliteti. "
        "Sinkronizirano s rums.json (app)."
    ])
    cols = [
        "Rum", "Kval /10", "Tip / Regija", "Aditiv status", "Izvor aditiv-podatka",
        "Cijena €", "Ducan", "Status", "Komentar",
        "Tijelo 1-5", "Slatkoća 1-5", "ABV %", "Pairable", "URL", "Serviranje",
    ]
    ws.append(cols)
    style_header(ws, 2, len(cols))

    current_tier = None
    for d in rums:
        q = float(d.get("qualityScore") or 0)
        tier = quality_tier_label(q)
        if tier != current_tier:
            ws.append([tier] + [None] * (len(cols) - 1))
            ws.cell(row=ws.max_row, column=1).font = SECTION_FONT
            current_tier = tier
        key = "|".join(sorted(match_tokens(d["name"])))
        status = d.get("status") or status_map.get(key) or ""
        ws.append([
            d["name"],
            q,
            d.get("region") or d.get("style") or "",
            additive_display(d),
            d.get("additiveSource") or "",
            price_str(d.get("priceEUR")),
            d.get("shopHR") or "",
            status,
            notes_hr(d),
            d.get("body"),
            d.get("sweetness"),
            d.get("abv") if d.get("abv") is not None else "",
            "✓" if d.get("pairable") else "",
            d.get("priceUrl") or "",
            (d.get("serving") or {}).get("best", ""),
        ])
    set_widths(ws, [36, 9, 18, 16, 22, 12, 16, 10, 55, 10, 12, 8, 9, 40, 22])
    ws.freeze_panes = "A3"

    # --- Po tipu ---
    idx = wb.sheetnames.index("Po tipu (kupnja)") if "Po tipu (kupnja)" in wb.sheetnames else None
    ws = replace_sheet(wb, "Po tipu (kupnja)", idx)
    ws.append([
        "PO TIPU / REGIJI — vodič za kupnju (1–2 iz grupe). Sinkronizirano s rums.json."
    ])
    pcols = ["Rum", "Kval /10", "Aditiv", "Cijena €", "Ducan/izvor", "Status / bilješka", "Tijelo", "URL"]
    ws.append(pcols)
    style_header(ws, 2, len(pcols))
    groups: dict[str, list] = defaultdict(list)
    for d in rums:
        if not d.get("pairable", True):
            continue
        label = d.get("region") or d.get("style") or "Ostalo"
        groups[label].append(d)
    for label in sorted(groups.keys(), key=lambda s: s.casefold()):
        items = sorted(groups[label], key=lambda x: -(x.get("qualityScore") or 0))
        ws.append([f"{label}  ({len(items)})"] + [None] * (len(pcols) - 1))
        ws.cell(row=ws.max_row, column=1).font = SECTION_FONT
        for d in items[:25]:
            st = d.get("status") or ""
            note = notes_hr(d)
            bilj = f"[{st}] {note}".strip() if st else note
            ws.append([
                d["name"],
                d.get("qualityScore"),
                additive_display(d),
                price_str(d.get("priceEUR")),
                d.get("shopHR") or "",
                bilj[:120],
                d.get("body"),
                d.get("priceUrl") or "",
            ])
    set_widths(ws, [36, 9, 14, 12, 18, 50, 8, 40])
    ws.freeze_panes = "A3"

    # --- Serviranje + Cigare ---
    idx = wb.sheetnames.index("Serviranje + Cigare") if "Serviranje + Cigare" in wb.sheetnames else None
    ws = replace_sheet(wb, "Serviranje + Cigare", idx)
    ws.append(["KAKO PITI + CIGARA PAIRING (po rumu) — sinkronizirano s rums.json"])
    scols = [
        "Rum / Profil", "Cisto", "Kap vode", "On the rocks", "Soda/highball", "Cola",
        "Najbolji nacin", "Cigara koja pase",
    ]
    ws.append(scols)
    style_header(ws, 2, len(scols))
    for d in rums:
        if not d.get("pairable", True):
            continue
        s = d.get("serving") or {}
        hint = cigar_hint_hr(d)
        if not hint:
            hint = ", ".join(d.get("flavorTags") or [])
        ws.append([
            d["name"],
            serve_cell(s, "neat"),
            serve_cell(s, "water"),
            serve_cell(s, "rocks"),
            serve_cell(s, "highball"),
            serve_cell(s, "cola"),
            serve_cell(s, "best"),
            hint,
        ])
    set_widths(ws, [40, 8, 10, 12, 12, 8, 28, 55])
    ws.freeze_panes = "A3"

    # Katalog / Svi rumovi / Kolekcija / Vodic ostaju netaknuti
    wb.save(path)
    print(f"Rum_Kolekcija_Checklist.xlsx: MASTER {len(rums)} (Katalog/Svi/Kolekcija/Vodič sačuvani)")


# --------------------------------------------------------------------------- whisky


def refresh_whisky_checklist() -> None:
    """MASTER + Po tipu + Serviranje iz whiskies.json; Katalog/Svi iz raw kataloga."""
    path = OUT / "Whisky_Kolekcija_Checklist.xlsx"
    whiskies: list[dict] = load_json("whiskies.json")
    whiskies = sorted(whiskies, key=lambda d: (-(d.get("qualityScore") or 0), d.get("name", "")))
    catalog = []
    if WHISKY_RAW.exists():
        catalog = json.loads(WHISKY_RAW.read_text(encoding="utf-8"))

    # Sačuvaj Kolekcija + Vodic ako postoje
    saved_kolekcija = None
    saved_vodic = None
    if path.exists():
        old = openpyxl.load_workbook(path)
        if "Kolekcija (plan)" in old.sheetnames:
            saved_kolekcija = [
                [c.value for c in row]
                for row in old["Kolekcija (plan)"].iter_rows()
            ]
        if "Vodic (sazetak)" in old.sheetnames:
            saved_vodic = [
                [c.value for c in row]
                for row in old["Vodic (sazetak)"].iter_rows()
            ]
        old.close()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Katalog
    ws = wb.create_sheet("Katalog allez+ecuga")
    ws.append(["Katalog whisky — allez.hr + ecuga.com"])
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.append(["Naziv", "Cijena", "Web shop", "URL"])
    style_header(ws, 2, 4)
    for item in catalog:
        ws.append([
            item.get("name", ""),
            format_price_eur(item.get("price_eur")),
            item.get("shop", ""),
            item.get("url", ""),
        ])
    set_widths(ws, [55, 12, 14, 50])

    # Svi viskiji (rang) — raw katalog + heuristika, seed override kad postoji
    seeds_by_tok = {"|".join(sorted(match_tokens(w["name"]))): w for w in whiskies}

    def find_seed(name: str) -> dict | None:
        key = "|".join(sorted(match_tokens(name)))
        if key in seeds_by_tok:
            return seeds_by_tok[key]
        best, best_score = None, 0
        for seed in whiskies:
            score = token_overlap(name, seed["name"])
            if score > best_score:
                best, best_score = seed, score
        return best if best and best_score >= 3 else None

    all_rows = []
    for item in catalog:
        name = item["name"]
        seed = find_seed(name)
        style, region, body, sweetness, tags = detect_style_region(
            name, item.get("ecuga_category", "")
        )
        if seed:
            style = seed.get("style", style)
            region = normalize_region(str(seed.get("region", region)), style)
            body = seed.get("body", body)
            sweetness = seed.get("sweetness", sweetness)
            tags = seed.get("flavorTags", tags)
        expr = detect_expression_type(name)
        abv = seed.get("abv") if seed and seed.get("abv") is not None else extract_abv(name)
        coloring = detect_coloring(name, style, expr)
        filt = detect_filter(name)
        price = item.get("price_eur")
        if seed and seed.get("priceEUR") and price is None:
            price = (seed["priceEUR"] or {}).get("min")
        quality = estimate_quality(
            name, price, style, expr, abv,
            seed_score=seed.get("qualityScore") if seed else None,
        )
        note = notes_hr(seed) if seed else f"Heuristika — {style}, {expr}"
        all_rows.append({
            "name": name,
            "quality": quality,
            "region": region,
            "style": style,
            "body": body,
            "sweetness": sweetness,
            "coloring": coloring,
            "filter": filt,
            "expr": expr,
            "abv": abv,
            "price": price,
            "shop": item.get("shop", ""),
            "url": item.get("url", ""),
            "note": note,
            "pairable": is_pairable(expr, style, quality),
            "seed": seed,
            "additive": additive_status(coloring, expr),
        })
    all_rows.sort(key=lambda r: (-r["quality"], r["name"]))

    ws = wb.create_sheet("Svi viskiji (rang)")
    ws.append(["SVI VISKIJI RANGIRANI po kvaliteti (sipping uz cigaru) | allez.hr + ecuga.com"])
    ws.append(["#", "Whisky", "Kval /10", "Tip / Regija", "Bilješka", "Cijena", "Shop", "ABV %", "URL"])
    style_header(ws, 2, 9)
    for i, row in enumerate(all_rows, 1):
        ws.append([
            i, row["name"], row["quality"], row["region"], row["note"][:80],
            format_price_eur(row["price"]), row["shop"],
            row["abv"] if row["abv"] is not None else "",
            row["url"],
        ])
    set_widths(ws, [4, 50, 9, 24, 40, 12, 12, 8, 40])
    ws.freeze_panes = "A3"

    # MASTER iz app whiskies.json (kurirani set) + append stupci
    ws = wb.create_sheet("MASTER Ocjene")
    ws.append(["MASTER — REKALIBRIRANO za SIPPING UZ CIGARU. Sinkronizirano s whiskies.json."])
    mcols = [
        "Whisky", "Kval /10", "Tip / Regija", "Boja (E150)", "Filter", "Expression",
        "Cijena €", "Dućan", "Status", "Komentar",
        "Tijelo 1-5", "Slatkoća 1-5", "ABV %", "Aditiv", "Pairable", "URL", "Serviranje",
    ]
    ws.append(mcols)
    style_header(ws, 2, len(mcols))
    current_tier = None
    for d in whiskies:
        q = float(d.get("qualityScore") or 0)
        tier = quality_tier_label(q)
        if tier != current_tier:
            ws.append([tier] + [None] * (len(mcols) - 1))
            ws.cell(row=ws.max_row, column=1).font = SECTION_FONT
            current_tier = tier
        # Boja/Filter iz imena + additiveDetail
        expr = detect_expression_type(d["name"])
        coloring = detect_coloring(d["name"], d.get("style", ""), expr)
        filt = detect_filter(d["name"])
        detail = d.get("additiveDetail")
        if isinstance(detail, dict) and detail.get("hr"):
            # pokušaj izvući boju iz detail teksta
            hr = detail["hr"].lower()
            if "prirodn" in hr or "natural" in hr:
                coloring = "natural"
        ws.append([
            d["name"],
            q,
            f"{d.get('style', '')} / {d.get('region', '')}".strip(" /"),
            coloring,
            filt,
            expr,
            price_str(d.get("priceEUR")),
            d.get("shopHR") or "",
            d.get("status") or "",
            notes_hr(d),
            d.get("body"),
            d.get("sweetness"),
            d.get("abv") if d.get("abv") is not None else "",
            additive_display(d),
            "✓" if d.get("pairable") else "",
            d.get("priceUrl") or "",
            (d.get("serving") or {}).get("best", ""),
        ])
    set_widths(ws, [40, 9, 28, 12, 10, 12, 12, 14, 10, 50, 10, 12, 8, 28, 9, 40, 22])
    ws.freeze_panes = "A3"

    # Po tipu
    ws = wb.create_sheet("Po tipu (kupnja)")
    ws.append(["PO TIPU — vodič za kupnju (1–2 iz grupe). Sinkronizirano s whiskies.json."])
    pcols = ["Whisky", "Kval /10", "Expression", "Cijena €", "Dućan/izvor", "Status / bilješka", "ABV %", "URL"]
    ws.append(pcols)
    style_header(ws, 2, len(pcols))
    style_labels = {
        "speyside-sherry": "Speyside sherry",
        "speyside-fruity": "Speyside fruity",
        "islay-peated": "Islay peated",
        "highland": "Highland",
        "campbeltown": "Campbeltown",
        "blended-scotch": "Blended Scotch",
        "bourbon": "Bourbon",
        "tennessee": "Tennessee",
        "rye": "Rye",
        "irish-pot-still": "Irish pot still",
        "irish-blend": "Irish blend",
        "japanese": "Japanese",
        "canadian": "Canadian",
        "world": "World whisky",
    }
    groups = defaultdict(list)
    for d in whiskies:
        if not d.get("pairable", True):
            continue
        label = style_labels.get(d.get("style", ""), d.get("style") or "Ostalo")
        groups[label].append(d)
    for label in sorted(groups.keys()):
        items = sorted(groups[label], key=lambda x: -(x.get("qualityScore") or 0))
        ws.append([f"{label}  ({len(items)})"] + [None] * (len(pcols) - 1))
        ws.cell(row=ws.max_row, column=1).font = SECTION_FONT
        for d in items[:25]:
            expr = detect_expression_type(d["name"])
            note = notes_hr(d)
            st = d.get("status") or ""
            bilj = f"[{st}] {note}".strip() if st else note
            ws.append([
                d["name"], d.get("qualityScore"), expr,
                price_str(d.get("priceEUR")), d.get("shopHR") or "",
                bilj[:120],
                d.get("abv") if d.get("abv") is not None else "",
                d.get("priceUrl") or "",
            ])
    set_widths(ws, [40, 9, 12, 12, 16, 50, 8, 40])
    ws.freeze_panes = "A3"

    # Serviranje
    ws = wb.create_sheet("Serviranje + Cigare")
    ws.append(["SERVIRANJE + CIGARE — profili + boce iz whiskies.json"])
    scols = ["Profil / primjer", "Neat", "Voda", "Rocks", "Best", "Cigar hint", "ABV %"]
    ws.append(scols)
    style_header(ws, 2, len(scols))
    profiles = [
        ("Speyside sherry (Aberlour, GlenDronach)", "speyside-sherry"),
        ("Islay peated (Ardbeg, Laphroaig)", "islay-peated"),
        ("Bourbon (Buffalo Trace, Four Roses)", "bourbon"),
        ("Irish pot still (Redbreast, Powers)", "irish-pot-still"),
        ("Japanese (Nikka, Yamazaki)", "japanese"),
        ("Blended Scotch (JW Black, Chivas)", "blended-scotch"),
    ]
    for label, style in profiles:
        s = serving_for_style(style, 46, "single-malt")
        ws.append([
            label,
            SERVE_MARK.get(int(s["neat"]), "+"),
            SERVE_MARK.get(int(s["water"]), "+"),
            SERVE_MARK.get(int(s["rocks"]), "~"),
            s["best"],
            resolve_serve_hint(label, style, cigar_hint_for_style),
            "",
        ])
    for d in whiskies:
        if not d.get("pairable", True):
            continue
        s = d.get("serving") or serving_for_style(d.get("style", ""), d.get("abv"), detect_expression_type(d["name"]))
        hint = cigar_hint_hr(d) or resolve_serve_hint(d["name"], d.get("style", ""), cigar_hint_for_style)
        ws.append([
            d["name"],
            serve_cell(s, "neat"),
            serve_cell(s, "water"),
            serve_cell(s, "rocks"),
            serve_cell(s, "best"),
            hint,
            d.get("abv") if d.get("abv") is not None else "",
        ])
    set_widths(ws, [42, 8, 8, 8, 28, 45, 8])
    ws.freeze_panes = "A3"

    # Kolekcija — restore ili default
    ws = wb.create_sheet("Kolekcija (plan)")
    if saved_kolekcija:
        for row in saved_kolekcija:
            ws.append(row)
        if ws.max_row >= 2:
            style_header(ws, 2, max(len(r) for r in saved_kolekcija if r) if saved_kolekcija else 8)
    else:
        ws.append(["TIER plan — whisky kolekcija (po stilu, bez duplikata)"])
        ws.append(["✓", "Tier", "Stil meta", "Boca meta", "Profil", "Izvor cijene", "Moja ocjena", "Bilješke"])
        style_header(ws, 2, 8)

    ws = wb.create_sheet("Vodic (sazetak)")
    if saved_vodic:
        for row in saved_vodic:
            ws.append(row)
    else:
        for line in [
            "WHISKY VODIČ — sažetak za sipping uz cigaru",
            "",
            "E150a bojenje: blended Scotch često koristi karamel (E150). Single malt obično natural.",
            "Chill-filter: NCF = nije chill-filtered (više tijela, bolje uz jaču cigaru).",
            "Izvor MASTER: whiskies.json | Katalog: scrape-whisky-catalog.py",
        ]:
            ws.append([line])

    wb.save(path)
    print(
        f"Whisky_Kolekcija_Checklist.xlsx: MASTER {len(whiskies)}, "
        f"katalog {len(catalog)}, Svi {len(all_rows)}"
    )


# --------------------------------------------------------------------------- cigars


def _cigar_pairing_score(c: dict) -> float:
    """Heuristički rang za MASTER (nema qualityScore u JSON-u)."""
    score = 5.0
    if "HR" in (c.get("markets") or []):
        score += 1.5
    if c.get("availabilityHR"):
        score += 0.5
    if notes_hr(c):
        score += 0.8
    if c.get("priceEUR") is not None:
        score += 0.4
    vitolas = c.get("vitolas") or []
    score += min(1.2, 0.25 * len(vitolas))
    if c.get("profileEstimated"):
        score -= 0.6
    tags = c.get("flavorTags") or []
    score += min(0.8, 0.15 * len(tags))
    # blago preferiraj srednju snagu za pairing
    strength = c.get("strength") or 3
    score += {1: 0.1, 2: 0.3, 3: 0.5, 4: 0.3, 5: 0.1}.get(strength, 0)
    return round(min(10.0, score), 1)


def _vitola_summary(c: dict) -> str:
    parts = []
    for v in c.get("vitolas") or []:
        bit = f"{v.get('name', '?')} ({v.get('smokeTimeMin', '?')}min"
        if v.get("priceEUR") is not None:
            bit += f", {v['priceEUR']}€"
        bit += ")"
        parts.append(bit)
    return "; ".join(parts)


def build_cigar_checklist() -> None:
    cigars: list[dict] = load_json("cigars.json")
    brands: dict = load_json("brands.json")

    scored = []
    for c in cigars:
        scored.append({**c, "_score": _cigar_pairing_score(c)})
    scored.sort(key=lambda c: (-c["_score"], c.get("brand", ""), c.get("line", "")))

    # MASTER: HR + bolje dokumentirane linije (cap ~220)
    master = []
    seen = set()
    for c in scored:
        key = (c.get("brand"), c.get("line"))
        if key in seen:
            continue
        hr = "HR" in (c.get("markets") or [])
        rich = bool(notes_hr(c)) and not c.get("profileEstimated")
        if hr or rich or c["_score"] >= 7.5:
            seen.add(key)
            master.append(c)
        if len(master) >= 220:
            break
    # ako i dalje malo HR, dodaj preostale HR
    if sum(1 for c in master if "HR" in (c.get("markets") or [])) < 80:
        for c in scored:
            key = (c.get("brand"), c.get("line"))
            if key in seen:
                continue
            if "HR" in (c.get("markets") or []):
                seen.add(key)
                master.append(c)
            if len(master) >= 280:
                break
    master.sort(key=lambda c: (-c["_score"], c.get("brand", ""), c.get("line", "")))

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # MASTER
    ws = wb.create_sheet("MASTER Ocjene")
    ws.append([
        "MASTER CIGARE — rang za pairing / kolekciju. "
        "Kval heuristika (HR dostupnost, bilješke, vitole). Sinkronizirano s cigars.json."
    ])
    mcols = [
        "Brand", "Linija", "Kval /10", "Zemlja", "Wrapper", "Format", "Vitola",
        "Snaga 1-5", "Tijelo 1-5", "Cijena €", "HR", "EU", "USA", "Svijet",
        "Vitole (⏱ min)", "Note", "URL", "Profil est.", "Komentar",
    ]
    ws.append(mcols)
    style_header(ws, 2, len(mcols))
    current = None
    for c in master:
        s = c.get("strength") or 3
        tier = (
            "Jake (4–5)" if s >= 4
            else "Srednje (3)" if s == 3
            else "Blaže (1–2)"
        )
        if tier != current:
            ws.append([tier] + [None] * (len(mcols) - 1))
            ws.cell(row=ws.max_row, column=1).font = SECTION_FONT
            current = tier
        m = c.get("markets") or []
        ws.append([
            c.get("brand", ""),
            c.get("line", ""),
            c["_score"],
            c.get("country", ""),
            c.get("wrapper", ""),
            c.get("format", ""),
            c.get("vitola", ""),
            c.get("strength"),
            c.get("body"),
            price_str(c.get("priceEUR")),
            "✓" if "HR" in m else "",
            "✓" if "EU" in m else "",
            "✓" if "USA" in m else "",
            "✓" if "WW" in m else "",
            _vitola_summary(c),
            ", ".join(c.get("flavorTags") or []),
            c.get("priceUrl") or "",
            "da" if c.get("profileEstimated") else "",
            notes_hr(c),
        ])
    set_widths(ws, [16, 24, 9, 14, 18, 12, 14, 9, 9, 10, 5, 5, 6, 7, 45, 28, 36, 10, 50])
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(mcols))}{ws.max_row}"

    # Svi cigare
    ws = wb.create_sheet("Svi cigare (rang)")
    ws.append(["SVI CIGARE RANGIRANI po pairing heuristici | cigars.json"])
    scols = [
        "#", "Brand", "Linija", "Kval /10", "Zemlja", "Wrapper", "Snaga", "Tijelo",
        "Cijena €", "HR", "Note", "Komentar",
    ]
    ws.append(scols)
    style_header(ws, 2, len(scols))
    for i, c in enumerate(scored, 1):
        m = c.get("markets") or []
        ws.append([
            i, c.get("brand", ""), c.get("line", ""), c["_score"],
            c.get("country", ""), c.get("wrapper", ""),
            c.get("strength"), c.get("body"),
            price_str(c.get("priceEUR")),
            "✓" if "HR" in m else "",
            ", ".join(c.get("flavorTags") or [])[:60],
            notes_hr(c)[:100],
        ])
    set_widths(ws, [5, 16, 24, 9, 14, 18, 8, 8, 10, 5, 28, 45])
    ws.freeze_panes = "A3"

    # Po tipu — zemlja
    ws = wb.create_sheet("Po tipu (kupnja)")
    ws.append([
        "PO ZEMLJI / WRAPPERU — vodič za kupnju (1–2 iz grupe). Unutar grupe po Kval."
    ])
    pcols = ["Brand / Linija", "Kval /10", "Wrapper", "Snaga", "Cijena €", "HR", "Status / bilješka"]
    ws.append(pcols)
    style_header(ws, 2, len(pcols))
    by_country: dict[str, list] = defaultdict(list)
    for c in master:
        by_country[c.get("country") or "Ostalo"].append(c)
    for country in sorted(by_country.keys(), key=lambda s: s.casefold()):
        items = sorted(by_country[country], key=lambda x: -x["_score"])
        ws.append([f"{country}  ({len(items)})"] + [None] * (len(pcols) - 1))
        ws.cell(row=ws.max_row, column=1).font = SECTION_FONT
        for c in items[:20]:
            m = c.get("markets") or []
            ws.append([
                f"{c.get('brand', '')} — {c.get('line', '')}",
                c["_score"],
                c.get("wrapper", ""),
                c.get("strength"),
                price_str(c.get("priceEUR")),
                "✓" if "HR" in m else "",
                notes_hr(c)[:100],
            ])
    set_widths(ws, [36, 9, 18, 8, 10, 5, 50])
    ws.freeze_panes = "A3"

    # Kolekcija plan
    ws = wb.create_sheet("Kolekcija (plan)")
    ws.append(["PLAN KOLEKCIJE — širina po stilovima wrappera / zemlje (checklist nabave)"])
    ws.append(["✓", "Razina", "Stil / fokus", "Cigara (meta)", "Profil", "Izvor / tržište", "Moja ocjena", "Bilješke"])
    style_header(ws, 2, 8)
    tiers = [
        ("S", "Connecticut (blago)", "Ashton Classic / Macanudo Cafe", "Kremasto, orasi, ulaz", "HR / EU", ""),
        ("S", "Habano srednje", "Partagas Serie / Romeo", "Začinsko, cedar", "HR / EU", ""),
        ("S", "Nikaragva core", "Padron 2000 / Oliva Serie V", "Čokolada, zemlja, snaga", "HR / EU", ""),
        ("A", "Maduro / San Andrés", "My Father / Oliva V Maduro", "Slatko-tamno, kakao", "HR / EU", ""),
        ("A", "Cuban klasik", "Montecristo No.2 / Partagas No.4", "Aromatska kompleksnost", "EU (put)", ""),
        ("A", "Corojo / pikantno", "My Father Le Bijou / Illusione", "Paprika, snaga", "HR / EU", ""),
        ("B", "Honduras", "Camacho / Punch", "Zemljano, robusno", "HR / EU", ""),
        ("B", "Connecticut broadleaf", "Liga Privada / Undercrown", "Ulje, kakao", "EU / USA", ""),
        ("B", "Sumatra / Ecuador", "Davidoff / Avo", "Elegantno, začin", "HR / EU", ""),
        ("C", "Strong box-pressed", "Liga No.9 / Padron 1926", "Punija večer", "EU", ""),
        ("C", "Short smoke", "Petites / robusto short", "30–45 min", "HR", ""),
        ("C", "Long smoke", "Churchill / Double Corona", "90+ min", "EU", ""),
    ]
    for tier, stil, meta, profil, izvor, bilj in tiers:
        ws.append(["", tier, stil, meta, profil, izvor, "", bilj])
    set_widths(ws, [4, 8, 24, 36, 28, 14, 12, 30])
    ws.freeze_panes = "A3"

    # Dim + Pairing (analog Serviranje + Cigare)
    ws = wb.create_sheet("Dim + Pairing")
    ws.append(["KAKO DIMITI + PIĆE PAIRING (po snazi / wrapperu)"])
    dcols = [
        "Profil / cigara", "Snaga", "Vrijeme (orijentir)", "Rez / draw", "Rum", "Whisky", "Best piće",
    ]
    ws.append(dcols)
    style_header(ws, 2, len(dcols))
    profiles = [
        ("Connecticut shade (blago)", "1–2", "45–70 min", "otvoren", "Agricole / suh Barbados", "Irish / Japanese", "Suhi rum ili Irish"),
        ("Ecuador Connecticut", "2", "50–75 min", "srednje", "Barbados XO", "Speyside fruity", "Barbados"),
        ("Habano / Cubanski stil", "3", "50–90 min", "srednje", "Kuba / Puerto Rico suh", "Speyside sherry", "Sherry malt ili kubanski rum"),
        ("Corojo / pikantno", "3–4", "50–80 min", "čvršći", "Jamajka esters", "Islay / rye", "Peated ili funk rum"),
        ("Maduro / San Andrés", "3–4", "60–100 min", "uljniji", "Demerara / solera", "Bourbon / sherry", "Bourbon ili Zacapa-stil"),
        ("Oscuro / broadleaf", "4–5", "70–120 min", "pun", "Navy / high ester", "Islay CS / rye", "Jaki whisky ili navy rum"),
        ("Nikaragva full body", "4", "60–100 min", "pun", "Demerara", "Bourbon SiB", "Bourbon"),
        ("Cuban mild-medium", "2–3", "45–75 min", "elegantno", "Santiago de Cuba / Eminente", "Cognac VSOP", "Cognac ili suh rum"),
    ]
    for row in profiles:
        ws.append(list(row))
    # top MASTER samples
    ws.append([])
    ws.append(["Primjeri iz MASTER seta"] + [None] * (len(dcols) - 1))
    ws.cell(row=ws.max_row, column=1).font = SECTION_FONT
    for c in master[:40]:
        strength = c.get("strength") or 3
        rum_h = "suh Barbados / agricole" if strength <= 2 else ("Jamajka / Demerara" if strength >= 4 else "Barbados XO / Habano-friendly")
        wh_h = "Irish / Japanese" if strength <= 2 else ("Islay / bourbon" if strength >= 4 else "Speyside / bourbon")
        ws.append([
            f"{c.get('brand')} {c.get('line')}",
            strength,
            f"{c.get('smokeTimeMin', '?')} min",
            c.get("format") or "",
            rum_h,
            wh_h,
            "Rum" if strength >= 3 else "Whisky/rum",
        ])
    set_widths(ws, [36, 8, 16, 14, 28, 22, 22])
    ws.freeze_panes = "A3"

    # Brendovi
    ws = wb.create_sheet("Brendovi")
    ws.append(["INDEKS BRENDOVA — brands.json + cigars.json"])
    bcols = ["Brand", "Zemlja", "Founded", "# linija", "# vitola", "Najjeftinija €", "Blurb HR"]
    ws.append(bcols)
    style_header(ws, 2, len(bcols))
    by_brand: dict[str, list] = defaultdict(list)
    for c in cigars:
        by_brand[c["brand"]].append(c)
    for brand in sorted(by_brand.keys(), key=lambda s: s.casefold()):
        lines = by_brand[brand]
        info = brands.get(brand, {})
        vitola_n = sum(len(c.get("vitolas") or []) for c in lines)
        prices = []
        for c in lines:
            if c.get("priceEUR") is not None:
                prices.append(float(c["priceEUR"]))
            for v in c.get("vitolas") or []:
                if v.get("priceEUR") is not None:
                    prices.append(float(v["priceEUR"]))
        blurb = (info.get("blurb") or {}).get("hr", "")
        if len(blurb) > 140:
            blurb = blurb[:137] + "..."
        ws.append([
            brand,
            info.get("country", ""),
            info.get("founded", ""),
            len(lines),
            vitola_n,
            min(prices) if prices else "",
            blurb,
        ])
    set_widths(ws, [22, 16, 12, 10, 10, 14, 60])
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:G{ws.max_row}"

    # Vodič
    ws = wb.create_sheet("Vodic (sazetak)")
    for line in [
        "CIGARE VODIČ — sažetak za pairing s pićem",
        "",
        "Snaga 1–5: orijentir za tijelo dima, ne 'kvalitetu'. Blaže (1–2) vole Connecticut i suhi rum/Irish.",
        "Wrapper: Connecticut = krem; Habano = začin; Maduro/San Andrés = slatko-tamno; Corojo = paprika.",
        "HR stupac: dostupnost na HR tržištu (markets). EU/USA/Svijet za filtriranje nabave.",
        "Kval /10 u MASTER-u je heuristika (HR + bilješke + vitole), ne laboratorijska ocjena.",
        "Profil est. = da → aroma/snaga procijenjeni, ne iz kuriranog opisa.",
        "",
        "Pairing palac: maduro → bourbon/sherry/solera rum; Habano → Speyside/kubanski rum; Connecticut → agricole/Irish.",
        "",
        "Regeneriraj: python scripts/build-kolekcija-checklists.py --only cigar",
        "Indeks (jednostavniji): Cigare_Index.xlsx via export-indexes.py",
    ]:
        ws.append([line])

    path = OUT / "Cigare_Kolekcija_Checklist.xlsx"
    wb.save(path)
    print(
        f"Cigare_Kolekcija_Checklist.xlsx: MASTER {len(master)}, "
        f"Svi {len(scored)}, brendova {len(by_brand)}"
    )


# --------------------------------------------------------------------------- generic drinks


def _serve_keys_for(items: list[dict]) -> list[str]:
    """Odredi stupce serviranja prema poljima u JSON-u (bez 'best')."""
    preferred = ["neat", "water", "rocks", "tonic", "martini", "highball", "cola"]
    present: set[str] = set()
    for d in items:
        s = d.get("serving") or {}
        if isinstance(s, dict):
            present.update(k for k in s if k != "best")
    return [k for k in preferred if k in present]


SERVE_COL_HR = {
    "neat": "Cisto / Neat",
    "water": "Voda",
    "rocks": "Rocks",
    "tonic": "Tonic",
    "martini": "Martini",
    "highball": "Highball",
    "cola": "Cola",
}


DRINK_CHECKLISTS: list[dict] = [
    {
        "key": "brandy",
        "json": "brandies.json",
        "xlsx": "Konjak_Brandy_Kolekcija_Checklist.xlsx",
        "title": "KONJAK / BRANDY",
        "item_label": "Brendy",
        "group_by": "style",
        "style_labels": {
            "cognac-vs": "Cognac VS",
            "cognac-vsop": "Cognac VSOP",
            "cognac-xo": "Cognac XO",
            "armagnac": "Armagnac",
            "calvados": "Calvados",
            "brandy-de-jerez": "Brandy de Jerez",
            "brandy-spanish": "Spanish brandy",
            "brandy-greek": "Greek (Metaxa)",
            "brandy-italian": "Italian",
            "brandy-armenian": "Armenian",
            "brandy-german": "German",
            "vinjak": "HR vinjak",
            "grappa": "Grappa",
        },
        "tiers": [
            ("1", "Cognac VSOP", "Remy VSOP / Hennessy VSOP", "Uravnotežen sipper", "allez.hr"),
            ("1", "Brandy de Jerez", "Carlos I / Fundador", "Orah + karamela", "allez.hr"),
            ("2", "Cognac XO", "Hennessy XO / Remy XO", "Bogat uz Habano", "allez.hr"),
            ("2", "Armagnac", "Janneau VSOP / Dartigalongue", "Rustikalniji profil", "allez.hr"),
            ("3", "Calvados", "Boulard VSOP / XO", "Jabuka + blaža cigara", "allez.hr"),
            ("3", "HR vinjak", "Badel Vinjak / Maraska", "Lokalni klasik", "lokalno"),
            ("4", "Cognac XO+", "Camus XO / Delamain", "Luksuzni sipper", "allez.hr"),
        ],
        "guide": [
            "VS / VSOP / XO: stariji tier = glađi, bolji uz cigaru (VS više za koktel).",
            "Doslađivanje: Jerez solera i Metaxa često slađi; cognac/armagnac obično čistiji profil.",
            "Pairing: Connecticut → VS/VSOP; Habano/oscuro → XO, Armagnac, Jerez solera.",
        ],
        "raw_catalog": "brandy",
    },
    {
        "key": "wine",
        "json": "wines.json",
        "xlsx": "Vino_Kolekcija_Checklist.xlsx",
        "title": "VINO",
        "item_label": "Vino",
        "group_by": "style",
        "style_labels": {
            "red-full": "Crveno puno",
            "red-medium": "Crveno srednje",
            "white-fresh": "Bijelo svježe",
            "sparkling": "Pjenušavo",
            "dessert-wine": "Desertno",
            "port-ruby": "Porto ruby",
            "port-tawny": "Porto tawny",
            "sherry-dry": "Sherry suhi",
        },
        "tiers": [
            ("1", "Porto tawny", "Tawny 10 / 20 YO", "Orah + karamela uz maduro", "vinoteka"),
            ("1", "Crveno puno", "Ribera / Priorat / Dingač", "Tijelo uz Habano", "vinoteka"),
            ("2", "Sherry oloroso", "Oloroso / PX", "Orah uz connecticut→maduro", "vinoteka"),
            ("2", "Porto ruby", "LBV / Vintage", "Voće + snaga", "vinoteka"),
            ("3", "Pjenušavo", "Champagne / Franciacorta", "Kontrast uz blagu cigaru", "vinoteka"),
            ("3", "HR crveno", "Teran / Plavac", "Lokalni pairing", "HR"),
            ("4", "Tokaj / Sauternes", "Aszú / Sauternes", "Desert uz blagi maduro", "vinoteka"),
        ],
        "guide": [
            "Uz cigaru najčešće: porto, sherry, punija crvena, desertna vina.",
            "Svježa bijela i lagani pjenušci: samo uz najblaže Connecticut profile.",
            "Temperatura serviranja važnija nego kod jakih pića — ne pregrijavaj porto/sherry.",
        ],
    },
    {
        "key": "gin",
        "json": "gins.json",
        "xlsx": "Gin_Kolekcija_Checklist.xlsx",
        "title": "GIN",
        "item_label": "Gin",
        "group_by": "style",
        "style_labels": {
            "london-dry": "London Dry",
            "plymouth": "Plymouth",
            "contemporary": "Contemporary",
            "premium-dry": "Premium dry",
            "croatian": "Hrvatski",
        },
        "tiers": [
            ("1", "London Dry klasik", "Beefeater / Tanqueray", "Juniper baza", "allez.hr"),
            ("1", "Premium dry", "Monkey 47 / Sipsmith", "Botanička širina", "allez.hr"),
            ("2", "Contemporary", "Roku / The Botanist", "Cvjetno / azijsko", "allez.hr"),
            ("2", "HR gin", "Local craft", "Lokalni terroir", "HR"),
            ("3", "Navy / high ABV", "Overproof stil", "Uz jaču cigaru", "allez.hr"),
        ],
        "guide": [
            "Uz cigaru: čisti / martini češće od G&T (tonic razvodni pairing).",
            "London Dry + Connecticut; contemporary citrus → blagi Habano; začinski gin → maduro oprezno.",
            "BotanicalProfile u Note stupcu pomaže birati kontraste (citrus vs zemlja).",
        ],
    },
    {
        "key": "coffee",
        "json": "coffees.json",
        "xlsx": "Kava_Kolekcija_Checklist.xlsx",
        "title": "KAVA",
        "item_label": "Kava",
        "group_by": "style",
        "style_labels": {
            "espresso-dark": "Espresso dark",
            "espresso-medium": "Espresso medium",
            "filter-light": "Filter light",
            "filter-medium": "Filter medium",
            "filter-dark": "Filter dark",
            "milk": "Mliječno",
            "spiked": "S alkoholom",
            "cold": "Cold brew / iced",
        },
        "tiers": [
            ("1", "Espresso medium-dark", "Italijanski blend", "Kakao + crema uz maduro", "Cogito / Quahwa"),
            ("1", "Filter medium", "Kenija / LatAm", "Uravnotežen uz Habano", "specialty"),
            ("2", "Filter light", "Etiopija Yirgacheffe", "Samo uz Connecticut", "specialty"),
            ("2", "Mliječno", "Cappuccino / flat white", "Krem uz blagu cigaru", "kafić"),
            ("3", "Cold brew", "Cold / nitro", "Ljetni pairing", "specialty"),
            ("3", "Spiked", "Irish / rum coffee", "Večernji desertni", "bar"),
        ],
        "guide": [
            "Svijetli filter (light) lako izgubi od dima — drži se Connecticut / blagih Habano.",
            "Espresso dark i mliječni napitci nose maduro i punije Nikaragva profile.",
            "Spiked coffee: biraj rum/whisky koji već radi s tom cigarom.",
        ],
    },
    {
        "key": "tequila",
        "json": "tequilas.json",
        "xlsx": "Tequila_Kolekcija_Checklist.xlsx",
        "title": "TEQUILA / MEZCAL",
        "item_label": "Tequila",
        "group_by": "style",
        "style_labels": {
            "blanco": "Blanco",
            "reposado": "Reposado",
            "anejo": "Añejo",
            "extra-anejo": "Extra Añejo",
        },
        "tiers": [
            ("1", "Reposado sipper", "Fortaleza / Ocho Reposado", "Agava + hrast uz Habano", "allez.hr"),
            ("1", "Blanco kvalitetan", "Tapatio / Fortaleza Blanco", "Čista agava uz Connecticut", "allez.hr"),
            ("2", "Añejo", "Añejo 18–36 mj", "Vanilija uz maduro", "allez.hr"),
            ("3", "Extra Añejo", "XA luksuz", "Teška večer", "allez.hr"),
            ("3", "Mezcal (ako u setu)", "Espadín / Tobalá", "Dim uz Islay-stil cigaru", "specijalizirano"),
        ],
        "guide": [
            "Blanco: čisto / ohlađeno; reposado/añejo: snifter, kao mladi brandy.",
            "Pairing: blanco → Connecticut/blagi Habano; añejo → maduro; mezcal dim → corojo/oscura.",
            "Izbjegavaj mixto i aromatizirane izraze u MASTER/app setu.",
        ],
    },
]


def _append_catalog_sheets(wb: openpyxl.Workbook, raw_path: Path, title: str) -> int:
    if not raw_path.exists():
        return 0
    catalog = json.loads(raw_path.read_text(encoding="utf-8"))
    ws = wb.create_sheet("Katalog allez+ecuga", 0)
    ws.append([f"Katalog {title} — allez.hr + ecuga.com"])
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.append(["Naziv", "Cijena", "Web shop", "URL"])
    style_header(ws, 2, 4)
    for item in catalog:
        ws.append([
            item.get("name", ""),
            format_price_eur(item.get("price_eur")),
            item.get("shop", ""),
            item.get("url", ""),
        ])
    set_widths(ws, [55, 12, 14, 50])
    return len(catalog)


def build_drink_checklist(cfg: dict) -> None:
    items: list[dict] = load_json(cfg["json"])
    items = sorted(items, key=lambda d: (-(d.get("qualityScore") or 0), d.get("name", "")))
    label = cfg["item_label"]
    title = cfg["title"]
    style_labels: dict = cfg.get("style_labels") or {}
    serve_keys = _serve_keys_for(items)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    catalog_n = 0
    if cfg.get("raw_catalog") == "brandy":
        catalog_n = _append_catalog_sheets(wb, BRANDY_RAW, "konjak/brandy")

    # MASTER
    ws = wb.create_sheet("MASTER Ocjene")
    ws.append([f"MASTER — {title} za sipping uz cigaru. Sinkronizirano s {cfg['json']}."])
    mcols = [
        label, "Kval /10", "Stil / Regija", "Tijelo 1-5", "Slatkoća 1-5",
        "Note", "Cijena €", "Dućan", "Status", "Komentar",
        "ABV %", "Aditiv", "Zemlja", "Pairable", "URL", "Serviranje",
    ]
    ws.append(mcols)
    style_header(ws, 2, len(mcols))
    current_tier = None
    for d in items:
        q = float(d.get("qualityScore") or 0)
        tier = quality_tier_label(q)
        if tier != current_tier:
            ws.append([tier] + [None] * (len(mcols) - 1))
            ws.cell(row=ws.max_row, column=1).font = SECTION_FONT
            current_tier = tier
        tags = d.get("flavorTags") or []
        if not tags and d.get("botanicalProfile"):
            bot = d["botanicalProfile"]
            tags = bot if isinstance(bot, list) else [str(bot)]
        ws.append([
            d.get("name", ""),
            q,
            f"{d.get('style', '')} / {d.get('region', '')}".strip(" /"),
            d.get("body"),
            d.get("sweetness"),
            ", ".join(tags),
            price_str(d.get("priceEUR")),
            d.get("shopHR") or "",
            d.get("status") or "",
            notes_hr(d),
            d.get("abv") if d.get("abv") is not None else "",
            additive_display(d),
            d.get("country") or "",
            "✓" if d.get("pairable") else "",
            d.get("priceUrl") or "",
            (d.get("serving") or {}).get("best", ""),
        ])
    set_widths(ws, [36, 9, 28, 10, 12, 28, 12, 16, 10, 50, 8, 18, 14, 9, 40, 22])
    ws.freeze_panes = "A3"

    # Svi (rang) — isti set, kompaktnije
    sheet_svi = {
        "Brendy": "Svi brendiji (rang)",
        "Vino": "Sva vina (rang)",
        "Gin": "Svi ginovi (rang)",
        "Kava": "Sve kave (rang)",
        "Tequila": "Sve tequile (rang)",
    }.get(label, "Svi (rang)")
    ws = wb.create_sheet(sheet_svi)
    ws.append([f"SVI — {title} rangirano po kvaliteti | {cfg['json']}"])
    scols = ["#", label, "Kval /10", "Stil / Regija", "Bilješka", "Cijena", "Shop", "ABV %", "URL"]
    ws.append(scols)
    style_header(ws, 2, len(scols))
    for i, d in enumerate(items, 1):
        ws.append([
            i,
            d.get("name", ""),
            d.get("qualityScore"),
            f"{d.get('style', '')} / {d.get('region', '')}".strip(" /"),
            notes_hr(d)[:80],
            price_str(d.get("priceEUR")),
            d.get("shopHR") or "",
            d.get("abv") if d.get("abv") is not None else "",
            d.get("priceUrl") or "",
        ])
    set_widths(ws, [4, 40, 9, 28, 45, 12, 14, 8, 40])
    ws.freeze_panes = "A3"

    # Po tipu
    ws = wb.create_sheet("Po tipu (kupnja)")
    ws.append([f"PO TIPU — vodič za kupnju (1–2 iz grupe). {title}."])
    pcols = [label, "Kval /10", "Cijena €", "Dućan/izvor", "Status / bilješka", "ABV %", "URL"]
    ws.append(pcols)
    style_header(ws, 2, len(pcols))
    groups: dict[str, list] = defaultdict(list)
    group_key = cfg.get("group_by", "style")
    for d in items:
        if not d.get("pairable", True):
            continue
        raw = d.get(group_key) or d.get("style") or "Ostalo"
        groups[style_labels.get(raw, raw)].append(d)
    for gname in sorted(groups.keys(), key=lambda s: s.casefold()):
        glist = sorted(groups[gname], key=lambda x: -(x.get("qualityScore") or 0))
        ws.append([f"{gname}  ({len(glist)})"] + [None] * (len(pcols) - 1))
        ws.cell(row=ws.max_row, column=1).font = SECTION_FONT
        for d in glist[:25]:
            st = d.get("status") or ""
            note = notes_hr(d)
            bilj = f"[{st}] {note}".strip() if st else note
            ws.append([
                d.get("name", ""),
                d.get("qualityScore"),
                price_str(d.get("priceEUR")),
                d.get("shopHR") or "",
                bilj[:120],
                d.get("abv") if d.get("abv") is not None else "",
                d.get("priceUrl") or "",
            ])
    set_widths(ws, [40, 9, 12, 16, 50, 8, 40])
    ws.freeze_panes = "A3"

    # Serviranje + Cigare
    ws = wb.create_sheet("Serviranje + Cigare")
    ws.append([f"SERVIRANJE + CIGARE — {title}"])
    scols2 = ["Profil / primjer", *[SERVE_COL_HR.get(k, k) for k in serve_keys], "Best", "Cigar hint"]
    ws.append(scols2)
    style_header(ws, 2, len(scols2))
    for d in items:
        if not d.get("pairable", True):
            continue
        s = d.get("serving") or {}
        hint = cigar_hint_hr(d)
        if not hint:
            hint = ", ".join(d.get("flavorTags") or [])
        row = [d.get("name", "")]
        for k in serve_keys:
            row.append(serve_cell(s, k))
        row.append(serve_cell(s, "best") or (s.get("best") if isinstance(s, dict) else "") or "")
        row.append(hint)
        ws.append(row)
    widths = [40] + [10] * len(serve_keys) + [28, 50]
    set_widths(ws, widths)
    ws.freeze_panes = "A3"

    # Kolekcija plan
    ws = wb.create_sheet("Kolekcija (plan)")
    ws.append([f"TIER plan — {title} kolekcija (po stilu, bez duplikata)"])
    ws.append(["✓", "Tier", "Stil meta", "Boca meta", "Profil", "Izvor cijene", "Moja ocjena", "Bilješke"])
    style_header(ws, 2, 8)
    for tier, stil, boca, profil, izvor in cfg.get("tiers") or []:
        ws.append(["", tier, stil, boca, profil, izvor, "", ""])
    set_widths(ws, [4, 8, 22, 36, 28, 14, 12, 30])
    ws.freeze_panes = "A3"

    # Vodič
    ws = wb.create_sheet("Vodic (sazetak)")
    lines = [
        f"{title} VODIČ — sažetak za sipping uz cigaru",
        "",
        *cfg.get("guide", []),
        "",
        f"Regeneriraj: python scripts/build-kolekcija-checklists.py --only {cfg['key']}",
        f"Indeks: *Index.xlsx via export-indexes.py | izvor: {cfg['json']}",
    ]
    for line in lines:
        ws.append([line])

    path = OUT / cfg["xlsx"]
    wb.save(path)
    extra = f", katalog {catalog_n}" if catalog_n else ""
    print(f"{cfg['xlsx']}: MASTER {len(items)}{extra}")


def main() -> int:
    all_keys = ["rum", "whisky", "brandy", "wine", "gin", "coffee", "tequila", "cigar"]
    parser = argparse.ArgumentParser()
    parser.add_argument("--only", choices=all_keys, default=None)
    args = parser.parse_args()
    targets = [args.only] if args.only else all_keys

    if "rum" in targets:
        refresh_rum_checklist()
    if "whisky" in targets:
        refresh_whisky_checklist()
    drink_cfgs = {c["key"]: c for c in DRINK_CHECKLISTS}
    for key in ("brandy", "wine", "gin", "coffee", "tequila"):
        if key in targets:
            build_drink_checklist(drink_cfgs[key])
    if "cigar" in targets:
        build_cigar_checklist()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

