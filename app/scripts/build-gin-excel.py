# -*- coding: utf-8 -*-
"""Build Gin_Kolekcija_Checklist.xlsx from gin_catalog_raw.json + seed gins.json.

Pokretanje: python scripts/build-gin-excel.py
"""
from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from gin_shared import (
    cigar_hint_for_style,
    detect_style_region,
    estimate_quality,
    extract_abv,
    format_price_eur,
    is_pairable,
    match_tokens,
    serving_for_style,
    token_overlap,
)

ROOT = Path(__file__).resolve().parent.parent
RAW = Path(__file__).resolve().parent / "output" / "gin_catalog_raw.json"
SEED = Path(__file__).resolve().parent / "seed" / "gins_seed.json"
APP_JSON = ROOT / "src" / "data" / "gins.json"
XLSX = ROOT.parent / "Gin_Kolekcija_Checklist.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="2C211A")
HEADER_FONT = Font(bold=True, color="C9A35C")
TITLE_FONT = Font(bold=True, size=12, color="9C4433")

MASTER_CAP = 80


def style_header(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center")


def load_seeds() -> list[dict]:
    path = SEED if SEED.exists() else APP_JSON
    if not path.exists():
        return []
    return json.loads(path.read_text(encoding="utf-8-sig"))


def load_seed_by_tokens(seeds_list: list[dict]) -> dict[str, dict]:
    by_tokens: dict[str, dict] = {}
    for item in seeds_list:
        key = "|".join(sorted(match_tokens(item["name"])))
        by_tokens[key] = item
    return by_tokens


def find_seed(name: str, seeds: dict[str, dict]) -> dict | None:
    key = "|".join(sorted(match_tokens(name)))
    if key in seeds:
        return seeds[key]
    best, best_score = None, 0
    for seed in seeds.values():
        score = token_overlap(name, seed["name"])
        if score > best_score:
            best, best_score = seed, score
    return best if best and best_score >= 2 else None


def enrich_row(item: dict, seed: dict | None) -> dict:
    name = item["name"]
    style, region, body, sweetness, botanical, tags = detect_style_region(
        name, item.get("ecuga_category", "")
    )
    if seed:
        style = seed.get("style", style)
        region = seed.get("region", region)
        body = seed.get("body", body)
        sweetness = seed.get("sweetness", sweetness)
        botanical = seed.get("botanicalProfile", botanical)
        tags = seed.get("flavorTags", tags)
    abv = seed.get("abv") if seed else extract_abv(name)
    if abv is None:
        abv = 40.0
    price = item.get("price_eur")
    if seed and seed.get("priceEUR") and price is None:
        price = (seed["priceEUR"] or {}).get("min")
    quality = estimate_quality(
        name, price, style, botanical, abv,
        seed_score=seed.get("qualityScore") if seed else None,
    )
    note = (seed or {}).get("notes", {}).get("hr", "")
    if not note:
        note = f"Heuristika — {style}, {botanical}"
    return {
        "name": name,
        "quality": quality,
        "style": style,
        "region": region,
        "body": body,
        "sweetness": sweetness,
        "botanical": botanical,
        "tags": tags,
        "abv": abv,
        "price": price,
        "shop": item.get("shop", ""),
        "url": item.get("url", ""),
        "note": note,
        "pairable": is_pairable(name, style, quality),
        "seed": seed is not None,
        "status": (seed or {}).get("status") or ("META" if seed and seed.get("meta") else ""),
    }


def append_orphan_seeds(catalog: list[dict], seeds: dict[str, dict]) -> list[dict]:
    catalog_names = {item["name"] for item in catalog}
    out = list(catalog)
    for seed in seeds.values():
        if any(token_overlap(seed["name"], cn) >= 3 for cn in catalog_names):
            continue
        out.append({
            "name": seed["name"],
            "price_eur": (seed.get("priceEUR") or {}).get("min"),
            "shop": seed.get("shopHR", "lokalno"),
            "url": seed.get("priceUrl") or "",
            "source": "seed",
        })
    return out


def build_svi_rang(catalog: list[dict], seeds: dict[str, dict]) -> list[dict]:
    rows = []
    for item in catalog:
        seed = find_seed(item["name"], seeds)
        rows.append(enrich_row(item, seed))
    rows.sort(key=lambda r: (-r["quality"], r["name"]))
    return rows


def select_master(all_rows: list[dict], seeds: dict[str, dict]) -> list[dict]:
    seed_names = {s["name"] for s in seeds.values()}
    master: list[dict] = []
    seen: set[str] = set()

    for row in all_rows:
        if row["seed"] or any(token_overlap(row["name"], sn) >= 3 for sn in seed_names):
            key = row["name"][:60]
            if key not in seen and row["pairable"]:
                seen.add(key)
                master.append(row)

    for row in all_rows:
        if len(master) >= MASTER_CAP:
            break
        if row["quality"] >= 6.8 and row["pairable"]:
            key = row["name"][:60]
            if key not in seen:
                seen.add(key)
                master.append(row)

    master.sort(key=lambda r: (-r["quality"], r["name"]))
    return master


def group_by_type(rows: list[dict]) -> list[tuple[str, list[dict]]]:
    groups: dict[str, list[dict]] = defaultdict(list)
    style_labels = {
        "london-dry": "London Dry",
        "premium-dry": "Premium Dry",
        "plymouth": "Plymouth",
        "contemporary": "Contemporary",
        "croatian": "HR gin",
    }
    for row in rows:
        if not row["pairable"]:
            continue
        label = style_labels.get(row["style"], row["style"])
        groups[label].append(row)
    out = []
    for label in sorted(groups.keys()):
        items = sorted(groups[label], key=lambda r: -r["quality"])
        out.append((label, items))
    return out


def write_workbook(catalog: list[dict], seeds: dict[str, dict]) -> None:
    catalog = append_orphan_seeds(catalog, seeds)
    all_rows = build_svi_rang(catalog, seeds)
    master_rows = select_master(all_rows, seeds)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_k = wb.create_sheet("Katalog allez+ecuga")
    ws_k.append(["Katalog gin — allez.hr + ecuga.com"])
    ws_k.cell(row=1, column=1).font = TITLE_FONT
    ws_k.append(["Naziv", "Cijena", "Web shop", "URL"])
    style_header(ws_k, 2, 4)
    for item in catalog:
        ws_k.append([
            item["name"],
            format_price_eur(item.get("price_eur")),
            item.get("shop", ""),
            item.get("url", ""),
        ])

    ws_s = wb.create_sheet("Svi ginovi (rang)")
    ws_s.append(["SVI GINOVI RANGIRANI po kvaliteti (sipping uz cigaru) | allez.hr + ecuga.com"])
    ws_s.append(["#", "Gin", "Kval /10", "Regija", "Bilješka", "Cijena", "Shop"])
    style_header(ws_s, 2, 7)
    for i, row in enumerate(all_rows, 1):
        ws_s.append([
            i,
            row["name"],
            row["quality"],
            row["region"],
            row["note"][:80],
            format_price_eur(row["price"]),
            row["shop"],
        ])

    ws_m = wb.create_sheet("MASTER Ocjene")
    ws_m.append(["MASTER — REKALIBRIRANO za SIPPING UZ CIGARU. Rangirano po kvaliteti."])
    ws_m.append([
        "Gin", "Kval /10", "Regija", "Stil", "Botanical",
        "Tijelo", "Slatkoća", "Cijena €", "Dućan", "Status", "Komentar",
    ])
    style_header(ws_m, 2, 11)
    current_tier = None
    for row in master_rows:
        tier = (
            "VRH za cigaru (8.0-10)" if row["quality"] >= 8
            else "Odlican sipper (7-8)" if row["quality"] >= 7
            else "Value / solidan (6-7)"
        )
        if tier != current_tier:
            ws_m.append([tier] + [None] * 10)
            current_tier = tier
        ws_m.append([
            row["name"],
            row["quality"],
            row["region"],
            row["style"],
            row["botanical"],
            row["body"],
            row["sweetness"],
            format_price_eur(row["price"]),
            row["shop"],
            row.get("status") or ("META" if row["seed"] and not row.get("url") else ""),
            row["note"],
        ])

    ws_t = wb.create_sheet("Po tipu (kupnja)")
    ws_t.append(["PO TIPU — vodic za kupnju (1-2 iz grupe). Rangirano po kvaliteti unutar grupe."])
    ws_t.append(["Gin", "Kval /10", "Stil", "Cijena €", "Ducan/izvor", "Status / biljeska"])
    style_header(ws_t, 2, 6)
    for label, items in group_by_type(master_rows):
        ws_t.append([f"{label}  ({len(items)})", None, None, None, None, None])
        for row in items[:25]:
            ws_t.append([
                row["name"],
                row["quality"],
                row["style"],
                format_price_eur(row["price"]),
                row["shop"],
                row["note"][:100],
            ])

    ws_sv = wb.create_sheet("Serviranje + Cigare")
    ws_sv.append(["SERVIRANJE + CIGARE — profili stilova za gin"])
    ws_sv.append(["Profil / primjer", "Neat", "Tonic", "Martini", "Best", "Cigar hint"])
    style_header(ws_sv, 2, 6)
    serve_map = {3: "++", 2: "+", 1: "~", 0: "x"}
    profiles = [
        ("London Dry (Beefeater / Tanqueray)", "london-dry", "classic-juniper"),
        ("Premium Dry (No. 3 / Sipsmith)", "premium-dry", "classic-juniper"),
        ("Contemporary (Monkey 47 / Roku)", "contemporary", "botanical"),
        ("Mediterranean (Gin Mare / Malfy)", "contemporary", "mediterranean"),
        ("HR gin", "croatian", "botanical"),
        ("Plymouth", "plymouth", "classic-juniper"),
    ]
    for label, style, botanical in profiles:
        s = serving_for_style(style, botanical)
        ws_sv.append([
            label,
            serve_map.get(s.get("neat", 2), "+"),
            serve_map.get(s.get("tonic", 2), "+"),
            serve_map.get(s.get("martini", 2), "+"),
            s["best"],
            cigar_hint_for_style(style, botanical),
        ])
    for row in master_rows[:40]:
        s = serving_for_style(row["style"], row["botanical"])
        ws_sv.append([
            row["name"][:50],
            serve_map.get(s.get("neat", 2), "+"),
            serve_map.get(s.get("tonic", 2), "+"),
            serve_map.get(s.get("martini", 2), "+"),
            s["best"],
            cigar_hint_for_style(row["style"], row["botanical"]),
        ])

    ws_v = wb.create_sheet("Vodic (sazetak)")
    for line in [
        "GIN VODIČ — sipping uz cigaru",
        "",
        "Uz cigaru: čisti / martini češće od G&T (tonic razvodni pairing).",
        "London Dry + Connecticut; contemporary citrus → blagi Habano; začinski gin → maduro oprezno.",
        "Flavoured / pink / sloe / RTD: u Katalogu, NE u MASTER/app (pairable=false).",
        "",
        "Pipeline: scrape-gin-catalog.py → build-gin-excel.py → [MASTER] → excel-to-gin-json.py",
        "Izvori: allez.hr/shop/gin1 + ecuga.com",
    ]:
        ws_v.append([line])

    wb.save(XLSX)
    print(f"Wrote {XLSX.name}")
    print(f"  Katalog: {len(catalog)} | Svi: {len(all_rows)} | MASTER: {len(master_rows)}")


def main() -> int:
    if not RAW.exists():
        print(f"Missing {RAW} — run scrape-gin-catalog.py first")
        return 1
    catalog = json.loads(RAW.read_text(encoding="utf-8"))
    seeds = load_seed_by_tokens(load_seeds())
    write_workbook(catalog, seeds)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
