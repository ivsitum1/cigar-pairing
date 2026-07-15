# -*- coding: utf-8 -*-
"""Ekstrakcija Konjak_Brandy_Checklist.xlsx -> brandies.json za pairing aplikaciju.

Cita MASTER Ocjene, Serviranje + Cigare i Katalog allez+ecuga sheetove.

Pokretanje: python scripts/excel-to-brandy-json.py
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

from brandy_shared import (
    detect_age_tier,
    detect_category_type,
    detect_style_region,
    detect_sweetening,
    extract_abv,
    find_best_catalog_match,
    is_pairable,
    match_tokens,
    parse_price_eur,
    serving_for_style,
    slugify,
    sweetening_to_additive,
    token_overlap,
)
from whisky_shared import normalize_region

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT.parent / "Konjak_Brandy_Checklist.xlsx"
OUT = ROOT / "src" / "data" / "brandies.json"
SEED = Path(__file__).resolve().parent / "seed" / "brandies_seed.json"

SERVE_SCORE = {"++": 3, "+": 2, "~": 1, "x": 0}


def build_catalog_index(wb) -> list[dict]:
    index = []
    for row in wb["Katalog allez+ecuga"].iter_rows(min_row=3, values_only=True):
        name, price, shop, url = row[0], row[1], row[2], row[3]
        if not name or not url:
            continue
        index.append({
            "name": str(name),
            "url": str(url),
            "shop": str(shop or ""),
            "price_eur": parse_price_eur(price),
            "tokens": match_tokens(str(name)),
        })
    return index


def build_serve_index(wb) -> list[dict]:
    rows = []
    ws = wb["Serviranje + Cigare"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        name, neat, water, rocks, best, hint = row[0], row[1], row[2], row[3], row[4], row[5]
        if not name or not best:
            continue
        rows.append({
            "tokens": match_tokens(str(name)),
            "serving": {
                "neat": SERVE_SCORE.get(str(neat).strip(), None) if neat else None,
                "water": SERVE_SCORE.get(str(water).strip(), None) if water else None,
                "rocks": SERVE_SCORE.get(str(rocks).strip(), None) if rocks else None,
                "highball": 0,
                "cola": 0,
                "best": str(best),
            },
            "cigarHint": str(hint) if hint else None,
        })
    return rows


def country_from_region(region: str) -> str:
    if "Francuska" in region:
        return "Francuska"
    if "Spanjolska" in region or "Španjolska" in region:
        return "Španjolska"
    if "Hrvatska" in region:
        return "Hrvatska"
    if "Grčka" in region:
        return "Grčka"
    if "Italija" in region:
        return "Italija"
    if "Armenija" in region:
        return "Armenija"
    if "Njemačka" in region:
        return "Njemačka"
    return region.split(",")[-1].strip() if "," in region else region


def load_seed_notes() -> list[dict]:
    if not SEED.exists():
        return []
    return json.loads(SEED.read_text(encoding="utf-8-sig"))


def find_seed_notes(name: str, seeds: list[dict]) -> dict | None:
    best, best_score = None, 0
    for seed in seeds:
        score = token_overlap(name, seed["name"])
        if score > best_score:
            best, best_score = seed, score
    return best if best and best_score >= 2 else None


def extract_brandies(wb, seeds: list[dict]) -> list[dict]:
    catalog = build_catalog_index(wb)
    serve_rows = build_serve_index(wb)
    ws = wb["MASTER Ocjene"]
    items = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        name, quality, region_col = row[0], row[1], row[2]
        if not name or quality is None:
            continue
        name_str = str(name)
        if name_str.startswith("VRH") or name_str.startswith("Odlican") or name_str.startswith("Value"):
            continue

        category = str(row[3] or detect_category_type(name_str))
        age_tier = str(row[4] or detect_age_tier(name_str))
        sweetening = str(row[5] or detect_sweetening(name_str, category))
        price_raw = row[6]
        shop = str(row[7] or "")
        comment = str(row[9] or "")

        style, region, body, sweetness, tags = detect_style_region(name_str)
        region_display = normalize_region(str(region_col or region), style)

        abv = extract_abv(name_str) or 40.0
        additive = sweetening_to_additive(sweetening, category)
        quality_f = float(quality)

        my_tokens = match_tokens(name_str)
        best_match, best_score = None, 0
        for sr in serve_rows:
            overlap = len(my_tokens & sr["tokens"])
            if overlap > best_score:
                best_match, best_score = sr, overlap
        if best_match and best_score >= 2:
            serving = dict(best_match["serving"])
        else:
            serving = serving_for_style(style, abv, category)

        cat = find_best_catalog_match(name_str, catalog)
        price_eur = parse_price_eur(price_raw)
        price_url = cat["url"] if cat else None
        if price_eur is None and cat and cat.get("price_eur"):
            price_eur = cat["price_eur"]
        if cat and not shop:
            shop = cat.get("shop", "")

        seed_match = find_seed_notes(name_str, seeds)
        seed_notes = (seed_match or {}).get("notes") or {}
        hr_note = comment
        if hr_note.startswith("Heuristika") and seed_notes.get("hr"):
            hr_note = seed_notes["hr"]

        items.append({
            "id": "br-" + slugify(name_str),
            "category": "brandy",
            "name": name_str,
            "style": style,
            "region": region_display,
            "country": country_from_region(region_display),
            "abv": abv,
            "body": body,
            "sweetness": sweetness,
            "flavorTags": tags,
            "additiveStatus": additive,
            "additiveDetail": f"{sweetening} / {age_tier}",
            "additiveSource": "Excel MASTER",
            "qualityScore": quality_f,
            "priceEUR": {"min": price_eur, "max": price_eur} if price_eur else None,
            "priceApprox": price_eur is None,
            "shopHR": shop or "allez.hr",
            "status": str(row[8]).strip() if row[8] and str(row[8]).strip() not in ("-", "") else None,
            "pairable": is_pairable(category, style, quality_f),
            "serving": serving,
            "cigarHint": None,
            "priceUrl": price_url,
            "notes": {"hr": hr_note, "en": seed_notes.get("en", "")},
        })
    return items


def main() -> int:
    if not XLSX.exists():
        print(f"Missing {XLSX} — run build-brandy-excel.py first")
        return 1
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    seeds = load_seed_notes()
    brandies = extract_brandies(wb, seeds)
    wb.close()
    OUT.write_text(json.dumps(brandies, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"brandies.json: {len(brandies)} brendija")
    with_url = sum(1 for b in brandies if b.get("priceUrl"))
    print(f"  priceUrl: {with_url}/{len(brandies)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
