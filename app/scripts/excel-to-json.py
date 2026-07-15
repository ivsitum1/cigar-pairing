# -*- coding: utf-8 -*-
"""Ekstrakcija Rum_Kolekcija_Checklist.xlsx -> JSON za pairing aplikaciju.

Cita MASTER Ocjene, Serviranje + Cigare i Kolekcija (plan) sheetove i generira:
  - src/data/rums.json      (kurirani rumovi s pairing atributima)
  - src/data/shopping.json  (tier plan kolekcije + ducani + preporuke)

Pokretanje:  python scripts/excel-to-json.py
Ponovno pokreni nakon svake izmjene Excela — JSON se generira ispocetka.
Prije exporta iz Excela: python scripts/export-serve-corrections.py (izvor istine)
Popravi lokalni Excel: python scripts/fix-excel-data.py
"""
import json
import re
import unicodedata
from pathlib import Path

import openpyxl

from serve_shared import find_correction, load_corrections

ROOT = Path(__file__).resolve().parent.parent          # app/
XLSX = ROOT.parent / "Rum_Kolekcija_Checklist.xlsx"
OUT = ROOT / "src" / "data"

# ---------------------------------------------------------------- helpers

# radni markeri iz Excela koji ne pripadaju u aplikaciju
WORK_MARKERS = re.compile(
    r"\s*\(?REKALIBRIRANO\s*\+?\)?\s*|\s*\[?(?:IMAS|META|PROBAO)\]?\s*|^NOVO\.\s*",
    re.IGNORECASE,
)

# Excel komentari su pisani bez dijakritika — vrati ih (cesti izrazi, cijela rijec)
DIACRITICS = {
    "cist": "čist", "cisti": "čisti", "cista": "čista", "cistac": "čistač",
    "najcisci": "najčišći", "cisci": "čišći",
    "suho-vocni": "suho-voćni", "vocni": "voćni", "voce": "voće",
    "jaca": "jača", "jacu": "jaču", "jace": "jače", "jaci": "jači",
    "laganija": "laganija", "blaza": "blaža", "slade": "slađe",
    "zacinski": "začinski", "zacini": "začini",
    "vise": "više", "trosi": "troši", "trosis": "trošiš",
    "presjece": "presiječe", "presijece": "presiječe", "gorcinu": "gorčinu",
    "snazna": "snažna", "snaznu": "snažnu", "ozbiljnu": "ozbiljnu",
    "cesto": "često", "ducan": "trgovina", "ducani": "trgovine",
    "jamajcanski": "jamajčanski", "jamajcanac": "jamajčanac",
    "kupis": "kupiš", "kvaris": "kvariš", "gladi": "glađi", "gladak": "gladak",
    "sarza": "šarža", "przeni": "prženi", "susi": "suši",
    "pise": "piše", "namjerno": "namjerno", "moze": "može",
}
_DIA_RE = re.compile(r"\b(" + "|".join(map(re.escape, DIACRITICS)) + r")\b")


def polish(text: str) -> str:
    """Ocisti radne markere, vikanje velikim slovima i vrati dijakritike."""
    if not text:
        return text
    s = WORK_MARKERS.sub(" ", text)
    # ALL-CAPS rijeci (3+ slova) -> normalno pisanje (JACA -> jaca, MADURO -> maduro)
    s = re.sub(r"\b[A-ZČĆŽŠĐ]{3,}\b",
               lambda m: m.group(0).lower() if m.group(0) not in
               ("AOC", "ECS", "HLCF", "OFTD", "VSOP", "XO") else m.group(0), s)
    s = _DIA_RE.sub(lambda m: DIACRITICS[m.group(1)], s)
    s = re.sub(r"\s{2,}", " ", s).strip(" .-–")
    if s and s[0].islower():
        s = s[0].upper() + s[1:]
    return s


def slugify(text: str) -> str:
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    text = re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")
    return text


def parse_price(raw):
    """'163-304 €' -> {min,max}; '82,80' -> {min:82.8,max:82.8}; inace None."""
    if not raw:
        return None
    s = str(raw).replace("€", "").strip()
    nums = re.findall(r"\d+(?:[.,]\d+)?", s)
    if not nums:
        return None
    vals = [float(n.replace(",", ".")) for n in nums]
    return {"min": min(vals), "max": max(vals)}


# ---------------------------------------------------------------- style model
# Stil se izvodi iz kolone 'Tip / Regija'; daje default body/sweetness/tagove.
# body: 1 (lagano) - 5 (puno tijelo) | sweetness: 1 (suho) - 5 (desertno)
STYLES = [
    # (regex na Tip/Regija, style id, body, sweetness, flavorTags)
    (r"SPICED|aromatiz", "spiced", 2, 5, ["vanilija", "zacini"]),
    (r"NIJE RUM|liker", "liqueur", 1, 5, ["slatko"]),
    (r"MIXING|NIJE SIPPING", "mixing", 1, 2, []),
    (r"Jamajka", "jamaica", 4, 1, ["ester-funk", "tropsko-voce", "hrast"]),
    (r"Agricole", "agricole", 2, 1, ["travnato", "vegetalno", "citrus"]),
    (r"Barbados", "barbados", 3, 1, ["suho-voce", "hrast", "vanilija"]),
    (r"Kuba", "cuba", 3, 1, ["zacini", "duhan", "citrus"]),
    (r"Demerara", "demerara", 4, 3, ["melasa", "tamno-voce", "karamela"]),
    (r"solera", "solera", 3, 3, ["karamela", "vanilija", "hrast"]),
    (r"Nikaragva", "nicaragua-dry", 3, 1, ["hrast", "suho-voce"]),
    (r"Kolumbija", "colombia", 3, 2, ["suho-voce", "vanilija"]),
    (r"Sv\. Lucija", "st-lucia", 3, 1, ["vanilija", "duhan", "hrast"]),
    (r"Trinidad", "trinidad", 3, 2, ["karamela", "hrast"]),
    (r"Puerto Rico", "puerto-rico", 3, 1, ["hrast", "vanilija"]),
    (r"Venezuela", "venezuela", 3, 2, ["karamela", "suho-voce"]),
    (r"Dominikana", "dominican", 3, 3, ["karamela", "vanilija"]),
    (r"Navy", "navy", 4, 2, ["melasa", "dim", "zacini"]),
    (r"Blend", "blend", 3, 2, ["suho-voce", "hrast"]),
    (r"Panama", "panama", 3, 3, ["karamela", "vanilija"]),
    (r"Reunion|Mauricijus|Fiji|Japan|Meksiko|Paragvaj|El Salvador|Peru|Kostarika|Gvatemala|Bermuda|Danska",
     "other", 3, 3, ["karamela"]),
]

ADDITIVE_MAP = [
    (r"aromatiz|spiced|kokos|naranca|elixir|liker", "flavored"),
    (r"doslazen|slazen|sladak|dodatak secera|visok", "sweetened"),
    (r"umjeren", "moderate"),
    (r"blag|nizak \(sherry\)", "light"),
    (r"cist|bez adit|bez dod|aoc|~bez|vrlo nis|nizak|0", "clean"),
]

# spiced/mixing/likeri nisu za pairing s cigarom
NON_PAIRABLE = {"spiced", "liqueur", "mixing"}


def detect_style(region: str):
    for pattern, style, body, sweet, tags in STYLES:
        if re.search(pattern, region, re.IGNORECASE):
            return style, body, sweet, list(tags)
    return "other", 3, 2, []


def normalize_additive(raw: str) -> str:
    low = (raw or "").lower()
    for pattern, norm in ADDITIVE_MAP:
        if re.search(pattern, low):
            return norm
    return "unknown"


# ---------------------------------------------------------------- serviranje

SERVE_SCORE = {"++": 3, "+": 2, "~": 1, "x": 0}

# default serviranje po stilu (iz sheeta Serviranje + Cigare, redovi profila)
SERVING_DEFAULTS = {
    "purist": {"neat": 3, "water": 3, "rocks": 1, "highball": 0, "cola": 0, "best": "Cisto / kap vode"},
    "jamaica": {"neat": 3, "water": 3, "rocks": 1, "highball": 1, "cola": 0, "best": "Kap vode (otvara estere)"},
    "agricole": {"neat": 3, "water": 3, "rocks": 1, "highball": 2, "cola": 0, "best": "Cisto / Ti' Punch"},
    "rich": {"neat": 2, "water": 2, "rocks": 3, "highball": 1, "cola": 0, "best": "On the rocks (velika kocka)"},
    "sweet": {"neat": 2, "water": 3, "rocks": 3, "highball": 1, "cola": 0, "best": "Velika kocka leda ili kap vode"},
    "mix": {"neat": 1, "water": 0, "rocks": 2, "highball": 3, "cola": 3, "best": "Koktel / highball"},
}

STYLE_TO_SERVING = {
    "jamaica": "jamaica", "agricole": "agricole",
    "demerara": "rich", "solera": "rich", "panama": "rich", "dominican": "rich",
    "navy": "rich", "other": "rich",
    "spiced": "mix", "liqueur": "mix", "mixing": "mix",
}  # sve ostalo -> purist


def serving_for(style: str, additive: str):
    if additive in ("sweetened", "flavored") and style not in NON_PAIRABLE:
        key = "sweet"
    else:
        key = STYLE_TO_SERVING.get(style, "purist")
    return dict(SERVING_DEFAULTS[key])


def match_tokens(name: str) -> set:
    stop = {"the", "de", "of", "and", "rum", "ron", "estate", "reserve", "reserva", "yo"}
    toks = set(re.findall(r"[a-z0-9]+", unicodedata.normalize("NFKD", name.lower())
                          .encode("ascii", "ignore").decode()))
    return {t for t in toks if t not in stop and not t.isdigit()}


# ---------------------------------------------------------------- extraction

def build_catalog_index(wb):
    """Katalog allez+ecuga sheet -> [(tokens, url, cijena)] za povezivanje linkova."""
    index = []
    for row in wb["Katalog allez+ecuga"].iter_rows(min_row=3, values_only=True):
        name, price, shop, url = row[0], row[1], row[2], row[3]
        if not name or not url:
            continue
        index.append({"tokens": match_tokens(str(name)), "url": str(url),
                      "name": str(name)})
    return index


def find_catalog_url(name: str, catalog) -> str | None:
    """Najbolji match po broju zajednickih tokena; prag >= 2."""
    tokens = match_tokens(name)
    best, best_score = None, 0
    for entry in catalog:
        score = len(tokens & entry["tokens"])
        if score > best_score:
            best, best_score = entry, score
    return best["url"] if best and best_score >= 2 else None


def extract_rums(wb):
    ws = wb["MASTER Ocjene"]
    catalog = build_catalog_index(wb)
    corrections = load_corrections()
    # serviranje redovi za name-match
    serve_rows = []
    for row in wb["Serviranje + Cigare"].iter_rows(min_row=3, values_only=True):
        name = row[0]
        if not name or not row[1]:  # header/prazno
            continue
        serve_rows.append({
            "tokens": match_tokens(str(name)),
            "serving": {
                "neat": SERVE_SCORE.get(str(row[1]).strip(), None),
                "water": SERVE_SCORE.get(str(row[2]).strip(), None),
                "rocks": SERVE_SCORE.get(str(row[3]).strip(), None),
                "highball": SERVE_SCORE.get(str(row[4]).strip(), None),
                "cola": SERVE_SCORE.get(str(row[5]).strip(), None),
                "best": row[6],
            },
            "cigarHint": row[7],
        })

    rums = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        name, quality, region = row[0], row[1], row[2]
        if not name or quality is None:  # sekcijski headeri / prazni redovi
            continue
        region = str(region or "")
        style, body, sweetness, tags = detect_style(region)
        additive = normalize_additive(str(row[3] or ""))
        comment = str(row[8] or "")

        # korekcije iz komentara / imena
        text = f"{name} {comment}".lower()
        if re.search(r"overproof|cask strength|\b6[0-9]\s?%|navy strength", text):
            body = min(5, body + 1)
            if "snaga" not in tags:
                tags.append("overproof")
        if additive in ("sweetened", "flavored"):
            sweetness = max(sweetness, 4)
        elif additive in ("light", "moderate"):
            sweetness = max(sweetness, 3)

        # match sa serviranje sheetom (token overlap >= 2 ili potpuno ime)
        my_tokens = match_tokens(str(name))
        best_match, best_score = None, 0
        for sr in serve_rows:
            overlap = len(my_tokens & sr["tokens"])
            if overlap > best_score:
                best_match, best_score = sr, overlap
        matched = best_match if best_score >= 2 else None

        serving = serving_for(style, additive)
        corr = find_correction(str(name), corrections)
        if corr:
            score_map = {"++": 3, "+": 2, "~": 1, "x": 0}
            for key in ("neat", "water", "rocks", "highball", "cola"):
                if corr.get(key) is not None:
                    serving[key] = score_map.get(str(corr[key]).strip(), serving.get(key, 0))
            if corr.get("best"):
                serving["best"] = corr["best"]
        elif matched:
            for k, v in matched["serving"].items():
                if v is not None:
                    serving[k if k != "best" else "best"] = v

        rums.append({
            "id": "rum-" + slugify(str(name)),
            "category": "rum",
            "name": str(name),
            "style": style,
            "region": region,
            "body": body,
            "sweetness": sweetness,
            "flavorTags": tags,
            "additiveStatus": additive,
            "additiveDetail": str(row[3] or ""),
            "additiveSource": str(row[4] or ""),
            "qualityScore": float(quality),
            "priceEUR": parse_price(row[5]),
            "shopHR": str(row[6] or ""),
            "status": (str(row[7]).strip() if row[7] and str(row[7]).strip() not in ("-", "") else None),
            "pairable": style not in NON_PAIRABLE and float(quality) >= 4,
            "serving": serving,
            "cigarHint": None,
            "priceUrl": find_catalog_url(str(name), catalog),
            "notes": {"hr": polish(comment), "en": ""},
        })
    return rums


def extract_shopping(wb):
    tiers, current_tier = [], None
    for row in wb["Kolekcija (plan)"].iter_rows(min_row=3, values_only=True):
        first, level = row[0], row[1]
        if first and str(first).startswith("TIER"):
            continue
        if not level:
            continue
        tiers.append({
            "tier": str(level).strip(),
            "owned": str(first or "").strip() == "✓",
            "styleTarget": str(row[2] or ""),
            "bottleTarget": str(row[3] or ""),
            "profile": polish(str(row[4] or "")),
            "priceSource": str(row[5] or "").replace("provjeriti", "").strip(" /"),
            "myRating": row[6],
            "notes": polish(str(row[7] or "")),
        })

    shops = [
        {"name": "allez.hr", "location": "Nova Ves 19, Zagreb / online", "note": {"hr": "Primarni HR izvor za rum i whisky, osobno preuzimanje", "en": "Primary Croatian source for rum & whisky, in-store pickup"}},
        {"name": "ecuga.com", "location": "online", "note": {"hr": "Sirok katalog rumova i zestica", "en": "Wide catalogue of rums and spirits"}},
        {"name": "Miva Galerija Vina", "location": "Zagreb", "note": {"hr": "Eminente, Dictador, Diplomatico", "en": "Eminente, Dictador, Diplomatico"}},
        {"name": "Cooltura To Go", "location": "Zagreb", "note": {"hr": "Botran linija (cesto rasprodano)", "en": "Botran range (often sold out)"}},
        {"name": "Roto Svijet Pica", "location": "Tkalciceva / Planinska, Zagreb", "note": {"hr": "Siroka ponuda", "en": "Wide selection"}},
        {"name": "Lidl", "location": "HR", "note": {"hr": "Planteray, Havana Club, Dos Maderas, La Hechicera — akcije", "en": "Planteray, Havana Club, Dos Maderas, La Hechicera — good deals"}},
        {"name": "Vivat Finavina", "location": "Zagreb", "note": {"hr": "Vise dessert/spiced; cisti: El Dorado 12, Pampero", "en": "Leans dessert/spiced; clean picks: El Dorado 12, Pampero"}},
        {"name": "Vrutak", "location": "Zagreb", "note": {"hr": "Premium/limited izdanja", "en": "Premium/limited releases"}},
        {"name": "Havana Shop (Camelot)", "location": "Zagreb, Split, Rovinj (+ Ljubljana)", "note": {"hr": "Ekskluzivni uvoznik Habanos i Davidoff cigara za HR/SLO — havana-cigar-shop.com", "en": "Exclusive Habanos & Davidoff cigar importer for Croatia/Slovenia — havana-cigar-shop.com"}},
        {"name": "The Humidor / Premium Cigars", "location": "Petrinjska 5 + Centar Kaptol, Zagreb", "note": {"hr": "Ista firma, 2 lokacije — New World boutique + kubanke, humidor.hr (cijene po vitoli)", "en": "Same company, 2 locations — New World boutique + Cubans, humidor.hr (per-vitola prices)"}},
        {"name": "Trafike / iCigara", "location": "HR", "note": {"hr": "VegaFina i cigarilosi u boljim trafikama; online prodaja duhana u HR nije dozvoljena", "en": "VegaFina and cigarillos in better tobacco kiosks; online tobacco sales are not allowed in Croatia"}},
    ]

    recommendations = [
        {"title": {"hr": "Najviše kvalitete po euru", "en": "Best quality per euro"}, "pick": "La Hechicera Solera 21", "detail": {"hr": "40,48 € (Lidl) — čist, suho-voćni profil", "en": "€40.48 (Lidl) — clean, dry-fruity profile"}},
        {"title": {"hr": "Najbolji omjer cijene i užitka uz cigaru", "en": "Best value alongside a cigar"}, "pick": "Havana Club 7 / Eminente 7", "detail": {"hr": "30,89 € (Lidl) / 55,20 € (Miva) — suhi kubanski stil", "en": "€30.89 (Lidl) / €55.20 (Miva) — dry Cuban style"}},
        {"title": {"hr": "Pouzdana baza za koktele", "en": "Reliable cocktail base"}, "pick": "Planteray 3 Stars", "detail": {"hr": "13,79 € (Lidl) — vrhunske boce sačuvajte za čistu degustaciju", "en": "€13.79 (Lidl) — save the finest bottles for neat sipping"}},
        {"title": {"hr": "Preporuka za proširenje kolekcije", "en": "Next addition to the collection"}, "pick": "Hampden Estate 8", "detail": {"hr": "75,55 € (allez.hr) — jamajčanski esterski karakter uz puniju cigaru", "en": "€75.55 (allez.hr) — Jamaican ester character for a fuller cigar"}},
    ]

    return {"tiers": tiers, "shops": shops, "recommendations": recommendations,
            "miniPath": ["Doorly's XO", "Hampden 8", "Flor de Cana 12", "El Dorado 12", "Clement VSOP"]}


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    OUT.mkdir(parents=True, exist_ok=True)

    rums = extract_rums(wb)
    (OUT / "rums.json").write_text(
        json.dumps(rums, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"rums.json: {len(rums)} rumova")

    shopping = extract_shopping(wb)
    (OUT / "shopping.json").write_text(
        json.dumps(shopping, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"shopping.json: {len(shopping['tiers'])} tier stavki, {len(shopping['shops'])} ducana")


if __name__ == "__main__":
    main()
