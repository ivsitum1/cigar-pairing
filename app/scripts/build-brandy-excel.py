# -*- coding: utf-8 -*-
"""Build Konjak_Brandy_Checklist.xlsx from brandy_catalog_raw.json + seed brandies.json.

Pokretanje: python scripts/build-brandy-excel.py
"""
from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from brandy_shared import (
    catalog_index,
    cigar_hint_for_style,
    detect_age_tier,
    detect_category_type,
    detect_style_region,
    detect_sweetening,
    estimate_quality,
    extract_abv,
    format_price_eur,
    is_pairable,
    match_tokens,
    serving_for_style,
    sweetening_to_additive,
    token_overlap,
)
from whisky_shared import normalize_region

ROOT = Path(__file__).resolve().parent.parent
RAW = Path(__file__).resolve().parent / "output" / "brandy_catalog_raw.json"
SEED = Path(__file__).resolve().parent / "seed" / "brandies_seed.json"
XLSX = ROOT.parent / "Konjak_Brandy_Checklist.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="2C211A")
HEADER_FONT = Font(bold=True, color="C9A35C")
TITLE_FONT = Font(bold=True, size=12, color="9C4433")

MASTER_CAP = 90


def style_header(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center")


def load_seed_by_tokens(seeds_list: list[dict]) -> dict[str, dict]:
    by_tokens: dict[str, dict] = {}
    for item in seeds_list:
        key = "|".join(sorted(match_tokens(item["name"])))
        by_tokens[key] = item
    return by_tokens


def find_seed(name: str, seeds: dict[str, dict]) -> dict | None:
    key = "|".join(sorted(match_tokens(name)))
    if key in seeds:
        return seeds[key]
    best, best_score = None, 0
    for seed in seeds.values():
        score = token_overlap(name, seed["name"])
        if score > best_score:
            best, best_score = seed, score
    return best if best and best_score >= 2 else None


def enrich_row(item: dict, seed: dict | None) -> dict:
    name = item["name"]
    style, region, body, sweetness, tags = detect_style_region(
        name, item.get("ecuga_category", "")
    )
    if seed:
        style = seed.get("style", style)
        region = normalize_region(str(seed.get("region", region)), style)
        body = seed.get("body", body)
        sweetness = seed.get("sweetness", sweetness)
        tags = seed.get("flavorTags", tags)
    category = detect_category_type(name, item.get("ecuga_category", ""))
    age_tier = detect_age_tier(name)
    abv = seed.get("abv") if seed else extract_abv(name)
    sweetening = detect_sweetening(name, category)
    price = item.get("price_eur")
    if seed and seed.get("priceEUR"):
        p = seed["priceEUR"]
        if price is None:
            price = p.get("min")
    quality = estimate_quality(
        name, price, style, category, age_tier, abv,
        seed_score=seed.get("qualityScore") if seed else None,
    )
    note = (seed or {}).get("notes", {}).get("hr", "")
    if not note:
        note = f"Heuristika — {style}, {category}, {age_tier}"
    return {
        "name": name,
        "quality": quality,
        "style": style,
        "region": region,
        "body": body,
        "sweetness": sweetness,
        "tags": tags,
        "category": category,
        "age_tier": age_tier,
        "sweetening": sweetening,
        "additive": sweetening_to_additive(sweetening, category),
        "abv": abv,
        "price": price,
        "shop": item.get("shop", ""),
        "url": item.get("url", ""),
        "note": note,
        "pairable": is_pairable(category, style, quality),
        "seed": seed is not None,
    }


def append_orphan_seeds(catalog: list[dict], seeds: dict[str, dict]) -> list[dict]:
    """HR vinjak i lokalne marke koje nisu u shop katalogu."""
    catalog_names = {item["name"] for item in catalog}
    out = list(catalog)
    for seed in seeds.values():
        if any(token_overlap(seed["name"], cn) >= 4 for cn in catalog_names):
            continue
        out.append({
            "name": seed["name"],
            "price_eur": (seed.get("priceEUR") or {}).get("min"),
            "shop": seed.get("shopHR", "lokalno"),
            "url": seed.get("priceUrl") or "",
            "source": "seed",
        })
    return out


def build_svi_rang(catalog: list[dict], seeds: dict[str, dict]) -> list[dict]:
    rows = []
    for item in catalog:
        seed = find_seed(item["name"], seeds)
        rows.append(enrich_row(item, seed))
    rows.sort(key=lambda r: (-r["quality"], r["name"]))
    return rows


def select_master(all_rows: list[dict], seeds: dict[str, dict]) -> list[dict]:
    seed_names = {s["name"] for s in seeds.values()}
    master: list[dict] = []
    seen: set[str] = set()

    for row in all_rows:
        if row["seed"] or any(token_overlap(row["name"], sn) >= 4 for sn in seed_names):
            key = row["name"][:60]
            if key not in seen and row["pairable"]:
                seen.add(key)
                master.append(row)

    for row in all_rows:
        if len(master) >= MASTER_CAP:
            break
        if row["quality"] >= 7.0 and row["pairable"]:
            key = row["name"][:60]
            if key not in seen:
                seen.add(key)
                master.append(row)

    master.sort(key=lambda r: (-r["quality"], r["name"]))
    return master


def group_by_type(rows: list[dict]) -> list[tuple[str, list[dict]]]:
    groups: dict[str, list[dict]] = defaultdict(list)
    style_labels = {
        "cognac-vs": "Cognac VS",
        "cognac-vsop": "Cognac VSOP",
        "cognac-xo": "Cognac XO",
        "armagnac": "Armagnac",
        "calvados": "Calvados",
        "brandy-de-jerez": "Brandy de Jerez",
        "brandy-spanish": "Spanish brandy",
        "brandy-greek": "Greek (Metaxa)",
        "brandy-italian": "Italian",
        "brandy-armenian": "Armenian",
        "brandy-german": "German",
        "vinjak": "HR vinjak",
        "brandy-irish": "Irish",
    }
    for row in rows:
        if not row["pairable"]:
            continue
        label = style_labels.get(row["style"], row["style"])
        groups[label].append(row)
    out = []
    for label in sorted(groups.keys()):
        items = sorted(groups[label], key=lambda r: -r["quality"])
        out.append((label, items))
    return out


def write_workbook(catalog: list[dict], seeds: dict[str, dict]) -> None:
    catalog = append_orphan_seeds(catalog, seeds)
    all_rows = build_svi_rang(catalog, seeds)
    master_rows = select_master(all_rows, seeds)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_k = wb.create_sheet("Katalog allez+ecuga")
    ws_k.append(["Katalog konjak/brandy — allez.hr + ecuga.com"])
    ws_k.cell(row=1, column=1).font = TITLE_FONT
    ws_k.append(["Naziv", "Cijena", "Web shop", "URL"])
    style_header(ws_k, 2, 4)
    for item in catalog:
        ws_k.append([
            item["name"],
            format_price_eur(item.get("price_eur")),
            item.get("shop", ""),
            item.get("url", ""),
        ])

    ws_s = wb.create_sheet("Svi brendiji (rang)")
    ws_s.append([
        "SVI BRENDIJI RANGIRANI po kvaliteti (sipping uz cigaru) | allez.hr + ecuga.com",
    ])
    ws_s.append(["#", "Brendy", "Kval /10", "Regija", "Bilješka", "Cijena", "Shop"])
    style_header(ws_s, 2, 7)
    for i, row in enumerate(all_rows, 1):
        ws_s.append([
            i,
            row["name"],
            row["quality"],
            row["region"],
            row["note"][:80],
            format_price_eur(row["price"]),
            row["shop"],
        ])

    ws_m = wb.create_sheet("MASTER Ocjene")
    ws_m.append(["MASTER — REKALIBRIRANO za SIPPING UZ CIGARU. Rangirano po kvaliteti."])
    ws_m.append([
        "Brendy", "Kval /10", "Regija", "Kategorija", "Age tier",
        "Doslađivanje", "Cijena €", "Dućan", "Status", "Komentar",
    ])
    style_header(ws_m, 2, 10)
    current_tier = None
    for row in master_rows:
        tier = (
            "VRH za cigaru (8.0-10)" if row["quality"] >= 8
            else "Odlican sipper (7-8)" if row["quality"] >= 7
            else "Value / solidan (6-7)"
        )
        if tier != current_tier:
            ws_m.append([tier, None, None, None, None, None, None, None, None, None])
            current_tier = tier
        ws_m.append([
            row["name"],
            row["quality"],
            row["region"],
            row["category"],
            row["age_tier"],
            row["sweetening"],
            format_price_eur(row["price"]),
            row["shop"],
            "META" if row["seed"] else "",
            row["note"],
        ])

    ws_t = wb.create_sheet("Po tipu (kupnja)")
    ws_t.append([
        "PO TIPU — vodic za kupnju (1-2 iz grupe). Rangirano po kvaliteti unutar grupe.",
    ])
    ws_t.append([
        "Brendy", "Kval /10", "Age tier", "Cijena €", "Ducan/izvor", "Status / biljeska",
    ])
    style_header(ws_t, 2, 6)
    for label, items in group_by_type(master_rows):
        ws_t.append([f"{label}  ({len(items)})", None, None, None, None, None])
        for row in items[:25]:
            ws_t.append([
                row["name"],
                row["quality"],
                row["age_tier"],
                format_price_eur(row["price"]),
                row["shop"],
                row["note"][:100],
            ])

    ws_sv = wb.create_sheet("Serviranje + Cigare")
    ws_sv.append(["SERVIRANJE + CIGARE — profili stilova za konjak/brandy"])
    ws_sv.append([
        "Profil / primjer", "Neat", "Voda", "Rocks", "Best", "Cigar hint",
    ])
    style_header(ws_sv, 2, 6)
    profiles = [
        ("Cognac VS (Hennessy VS)", "cognac-vs", "cognac"),
        ("Cognac VSOP (Remy VSOP)", "cognac-vsop", "cognac"),
        ("Cognac XO (Hennessy XO)", "cognac-xo", "cognac"),
        ("Armagnac (Janéau VSOP)", "armagnac", "armagnac"),
        ("Brandy de Jerez (Carlos I)", "brandy-de-jerez", "brandy-jerez"),
        ("Calvados (Boulard)", "calvados", "calvados"),
        ("HR vinjak (Badel)", "vinjak", "vinjak"),
    ]
    serve_map = {"3": "++", "2": "+", "1": "~", "0": "x"}
    for label, style, cat in profiles:
        s = serving_for_style(style, 40, cat)
        ws_sv.append([
            label,
            serve_map.get(str(s["neat"]), "+"),
            serve_map.get(str(s["water"]), "+"),
            serve_map.get(str(s["rocks"]), "~"),
            s["best"],
            cigar_hint_for_style(style),
        ])
    for row in master_rows[:35]:
        s = serving_for_style(row["style"], row.get("abv"), row["category"])
        ws_sv.append([
            row["name"],
            serve_map.get(str(s["neat"]), "+"),
            serve_map.get(str(s["water"]), "+"),
            serve_map.get(str(s["rocks"]), "~"),
            s["best"],
            cigar_hint_for_style(row["style"]),
        ])

    ws_p = wb.create_sheet("Kolekcija (plan)")
    ws_p.append(["TIER plan — konjak/brandy kolekcija (po stilu, bez duplikata)"])
    ws_p.append([
        "✓", "Tier", "Stil meta", "Boca meta", "Profil", "Izvor cijene", "Moja ocjena", "Biljeske",
    ])
    style_header(ws_p, 2, 8)
    tiers = [
        ("TIER 1 — must have", "Cognac VSOP", "Remy VSOP / Hennessy VSOP", "Uravnotezen sipper", "allez.hr", "7+"),
        ("TIER 1 — must have", "Brandy de Jerez", "Carlos I / Fundador", "Orah + karamela", "allez.hr", "8+"),
        ("TIER 2 — core", "Cognac XO", "Hennessy XO / Remy XO", "Bogat uz Habano", "allez.hr", "8+"),
        ("TIER 2 — core", "Armagnac", "Janéau VSOP / Dartigalongue", "Rustikalniji profil", "allez.hr", "7.5+"),
        ("TIER 3 — explore", "Calvados", "Boulard VSOP / XO", "Jabuka + blaza cigara", "allez.hr", "7+"),
        ("TIER 3 — explore", "HR vinjak", "Badel 5* / Stock XO", "Lokalni klasik", "lokalno", "7+"),
        ("TIER 4 — luxury", "Cognac XO+", "Camus XO / Delamain", "Luksuzni sipper", "allez.hr", "8.5+"),
    ]
    for t in tiers:
        ws_p.append(["", *t, "", ""])

    ws_g = wb.create_sheet("Vodic (sazetak)")
    guide = [
        "KONJAK / BRANDY VODIC — sazetak za sipping uz cigaru",
        "",
        "VS / VSOP / XO: stariji tier = glađi, bolji uz cigaru (VS vise za koktel).",
        "Dosladivanje: Jerez solera i Metaxa cesto sladji; cognac/armagnac obicno cistiji profil.",
        "Grappa / pisco / absinthe / likeri: u punom katalogu, NE u MASTER/app (pairable=false).",
        "Pairing: Connecticut -> VS/VSOP; Habano/oscuro -> XO, Armagnac, Jerez solera.",
        "",
        "Izvori cijena: allez.hr (primarni) + ecuga.com. Regeneriraj: scrape-brandy-catalog.py",
    ]
    for line in guide:
        ws_g.append([line])

    wb.save(XLSX)
    print(f"Wrote {XLSX}")
    print(f"  Katalog: {len(catalog)} | Svi: {len(all_rows)} | MASTER: {len(master_rows)}")


def main() -> int:
    if not RAW.exists():
        print(f"Missing {RAW} — run scrape-brandy-catalog.py first")
        return 1
    catalog = json.loads(RAW.read_text(encoding="utf-8"))
    seeds_list = json.loads(SEED.read_text(encoding="utf-8-sig")) if SEED.exists() else []
    seeds = load_seed_by_tokens(seeds_list)
    write_workbook(catalog, seeds)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
