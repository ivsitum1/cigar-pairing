# -*- coding: utf-8 -*-
"""Ekstrakcija Gin_Kolekcija_Checklist.xlsx -> gins.json.

Cita MASTER Ocjene, Serviranje + Cigare i Katalog allez+ecuga sheetove.
Zadrzava postojece unose iz gins.json koji nisu u MASTER-u (merge).

Pokretanje: python scripts/excel-to-gin-json.py
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

from gin_shared import (
    cigar_hint_pair,
    detect_style_region,
    extract_abv,
    find_best_catalog_match,
    is_gift_sku,
    is_templated_hint,
    match_tokens,
    parse_price_eur,
    serving_for_style,
    slugify,
    token_overlap,
)

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT.parent / "Gin_Kolekcija_Checklist.xlsx"
OUT = ROOT / "src" / "data" / "gins.json"
SEED = Path(__file__).resolve().parent / "seed" / "gins_seed.json"
ALIASES = ROOT / "src" / "data" / "drinkIdAliases.json"

SERVE_SCORE = {"++": 3, "+": 2, "~": 1, "x": 0}


def build_catalog_index(wb) -> list[dict]:
    index = []
    if "Katalog allez+ecuga" not in wb.sheetnames:
        return index
    for row in wb["Katalog allez+ecuga"].iter_rows(min_row=3, values_only=True):
        name, price, shop, url = row[0], row[1], row[2], row[3]
        if not name or not url:
            continue
        index.append({
            "name": str(name),
            "url": str(url),
            "shop": str(shop or ""),
            "price_eur": parse_price_eur(price),
            "tokens": match_tokens(str(name)),
        })
    return index


def _is_profile_label(name: str) -> bool:
    text = name.strip().lower()
    return "(" in name or text.startswith((
        "london dry", "premium dry", "contemporary", "mediterranean", "hr gin", "plymouth",
    ))


def build_serve_index(wb) -> list[dict]:
    rows = []
    if "Serviranje + Cigare" not in wb.sheetnames:
        return rows
    ws = wb["Serviranje + Cigare"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        name, neat, tonic, martini, best, hint = (
            row[0], row[1], row[2], row[3], row[4], row[5] if len(row) > 5 else None
        )
        hint_en = row[6] if len(row) > 6 else None
        if not name or not best:
            continue
        name_str = str(name)
        rows.append({
            "name": name_str,
            "tokens": match_tokens(name_str),
            "gift": is_gift_sku(name_str),
            "profile": _is_profile_label(name_str),
            "serving": {
                "neat": SERVE_SCORE.get(str(neat).strip(), None) if neat else None,
                "tonic": SERVE_SCORE.get(str(tonic).strip(), None) if tonic else None,
                "martini": SERVE_SCORE.get(str(martini).strip(), None) if martini else None,
                "highball": 2,
                "best": str(best),
            },
            "cigarHintHr": str(hint).strip() if hint else None,
            "cigarHintEn": str(hint_en).strip() if hint_en else None,
        })
    return rows


def load_aliases() -> dict[str, str]:
    if not ALIASES.exists():
        return {}
    data = json.loads(ALIASES.read_text(encoding="utf-8-sig"))
    return data.get("aliases") or {}


def load_existing() -> list[dict]:
    if OUT.exists():
        return json.loads(OUT.read_text(encoding="utf-8-sig"))
    if SEED.exists():
        return json.loads(SEED.read_text(encoding="utf-8-sig"))
    return []


def find_existing(name: str, existing: list[dict]) -> dict | None:
    """Match a MASTER row to a live record without stealing a short sipping id."""
    target_gift = is_gift_sku(name)
    name_l = name.lower().strip()
    for item in existing:
        if (item.get("name") or "").lower().strip() == name_l:
            return item

    best, best_score = None, 0
    for item in existing:
        item_name = item.get("name") or ""
        if is_gift_sku(item_name) != target_gift:
            continue
        score = token_overlap(name, item_name)
        if score < 3:
            continue
        shorter = min(len(name), len(item_name))
        longer = max(len(name), len(item_name))
        # Short referent ("Sipsmith") must not attach to a much longer gift title.
        if longer >= shorter + 24 and score < 5:
            continue
        if score > best_score:
            best, best_score = item, score
    return best


def extract_gins(wb, existing: list[dict]) -> list[dict]:
    catalog = build_catalog_index(wb)
    serve_rows = build_serve_index(wb)
    aliases = load_aliases()
    existing_by_id = {item["id"]: item for item in existing if item.get("id")}
    ws = wb["MASTER Ocjene"]
    items = []
    seen_ids: set[str] = set()

    for row in ws.iter_rows(min_row=3, values_only=True):
        name, quality, region_col = row[0], row[1], row[2]
        if not name or quality is None:
            continue
        name_str = str(name)
        if name_str.startswith("VRH") or name_str.startswith("Odlican") or name_str.startswith("Value"):
            continue

        style_col = str(row[3] or "")
        botanical_col = str(row[4] or "")
        body_col = row[5]
        sweet_col = row[6]
        price_raw = row[7]
        shop = str(row[8] or "")
        status = str(row[9]).strip() if row[9] and str(row[9]).strip() not in ("-", "") else None
        comment = str(row[10] or "") if len(row) > 10 else ""

        style, region, body, sweetness, botanical, tags = detect_style_region(name_str)
        if style_col:
            style = style_col
        if botanical_col:
            botanical = botanical_col
        if region_col:
            region = str(region_col)
        if body_col is not None:
            body = int(body_col)
        if sweet_col is not None:
            sweetness = int(sweet_col)

        abv = extract_abv(name_str) or 40.0
        quality_f = float(quality)

        my_tokens = match_tokens(name_str)
        my_gift = is_gift_sku(name_str)
        best_match, best_score = None, 0
        for sr in serve_rows:
            if sr.get("gift") != my_gift:
                continue
            overlap = len(my_tokens & sr["tokens"])
            if overlap > best_score:
                best_match, best_score = sr, overlap
        excel_hint_hr = excel_hint_en = None
        if best_match and best_score >= 2:
            serving = {k: v for k, v in best_match["serving"].items() if v is not None}
            if "best" not in serving:
                serving["best"] = serving_for_style(style, botanical)["best"]
            if not best_match.get("profile"):
                excel_hint_hr = best_match.get("cigarHintHr")
                excel_hint_en = best_match.get("cigarHintEn")
        else:
            serving = serving_for_style(style, botanical)

        cat = find_best_catalog_match(name_str, catalog)
        if cat and is_gift_sku(cat.get("name") or "") != my_gift:
            cat = None
        price_eur = parse_price_eur(price_raw)
        if price_eur is None and cat and cat.get("price_eur"):
            price_eur = cat["price_eur"]
        if cat and not shop:
            shop = cat.get("shop", "")

        prev = find_existing(name_str, existing)
        seed_notes = (prev or {}).get("notes") or {}
        hr_note = comment
        if (not hr_note or hr_note.startswith("Heuristika")) and seed_notes.get("hr"):
            hr_note = seed_notes["hr"]

        cid = (prev or {}).get("id") or ("gin-" + slugify(name_str))
        if cid in aliases:
            cid = aliases[cid]
            prev = existing_by_id.get(cid, prev)
        display_name = name_str
        if prev and token_overlap(name_str, prev["name"]) >= 2 and len(prev["name"]) < len(name_str):
            if is_gift_sku(prev["name"]) == my_gift:
                display_name = prev["name"]
                cid = prev["id"]
                if prev.get("abv"):
                    abv = prev["abv"]
                if prev.get("flavorTags"):
                    tags = prev["flavorTags"]

        generated = cigar_hint_pair(display_name, style, botanical)
        cigar_hint = {
            "hr": generated["hr"],
            "en": generated["en"],
        }
        if excel_hint_hr and len(excel_hint_hr) >= 80:
            cigar_hint["hr"] = excel_hint_hr
            if excel_hint_en and len(excel_hint_en) >= 40:
                cigar_hint["en"] = excel_hint_en

        prev_url = (prev or {}).get("priceUrl")
        if cat:
            price_url = cat["url"]
        else:
            price_url = prev_url
        if (prev or {}).get("meta") and not prev_url:
            # META-only bottles absent from scrape must not grow invented shop URLs.
            price_url = None

        item = {
            "id": cid,
            "name": display_name,
            "style": style,
            "region": region,
            "botanicalProfile": botanical,
            "body": body,
            "sweetness": sweetness,
            "flavorTags": tags,
            "qualityScore": quality_f,
            "abv": abv,
            "priceEUR": {"min": price_eur, "max": price_eur} if price_eur else None,
            "shopHR": shop or "allez.hr",
            "priceUrl": price_url,
            "pairable": True,
            "serving": serving,
            "cigarHint": cigar_hint,
            "notes": {
                "hr": hr_note,
                "en": seed_notes.get("en", ""),
            },
            "category": "gin",
        }
        if prev and prev.get("profileEstimated") is not None:
            item["profileEstimated"] = prev["profileEstimated"]
        if status == "META" or (prev or {}).get("meta") or my_gift:
            item["meta"] = True
            item["status"] = "META"
        elif status:
            item["status"] = status

        if cid in seen_ids:
            continue
        seen_ids.add(cid)
        items.append(item)

    # Keep every live id. Token overlap must not drop a sipping referent.
    for prev in existing:
        if prev.get("id") in seen_ids:
            continue
        if prev.get("id") in aliases:
            continue
        prev = dict(prev)
        prev["pairable"] = True
        hint = prev.get("cigarHint") or {}
        if is_templated_hint(hint.get("hr")):
            prev["cigarHint"] = cigar_hint_pair(
                prev.get("name") or "",
                prev.get("style") or "",
                prev.get("botanicalProfile") or "",
            )
        if is_gift_sku(prev.get("name") or ""):
            prev["meta"] = True
            prev["status"] = "META"
        items.append(prev)
        seen_ids.add(prev["id"])

    items.sort(key=lambda d: (-(d.get("qualityScore") or 0), d.get("name", "")))
    return items


def main() -> int:
    if not XLSX.exists():
        print(f"Missing {XLSX} — run build-gin-excel.py first")
        return 1
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    existing = load_existing()
    gins = extract_gins(wb, existing)
    wb.close()
    OUT.write_text(json.dumps(gins, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"gins.json: {len(gins)} ginova")
    with_url = sum(1 for g in gins if g.get("priceUrl"))
    print(f"  priceUrl: {with_url}/{len(gins)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
