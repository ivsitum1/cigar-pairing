# -*- coding: utf-8 -*-
"""Generira Excel indekse po kategoriji iz app/src/data/*.json — istim stilom
kao Rum_Kolekcija_Checklist.xlsx (rangirano po kvaliteti za sipping uz cigaru).

Pokretanje: python scripts/export-indexes.py
Izlaz (u root folderu projekta):
  Whisky_Index.xlsx, Konjak_Brandy_Index.xlsx, Kava_Index.xlsx, Cigare_Index.xlsx
"""
import json
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parent.parent  # app/
DATA = ROOT / "src" / "data"
OUT = ROOT.parent  # cigar_and_rum/

HEADER_FILL = PatternFill("solid", fgColor="2C211A")
HEADER_FONT = Font(bold=True, color="C9A35C")
TITLE_FONT = Font(bold=True, size=12, color="9C4433")


def price_str(p):
    if p is None:
        return "provjeriti"
    if isinstance(p, (int, float)):
        return f"{p:.0f} €"
    if abs(p["min"] - p["max"]) < 0.01:
        return f"{p['min']:.2f} €"
    return f"{p['min']:.0f}-{p['max']:.0f} €"


def style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center")


def export_drinks(json_name, xlsx_name, title, sheet_title):
    items = json.loads((DATA / json_name).read_text(encoding="utf-8"))
    items.sort(key=lambda d: -(d.get("qualityScore") or 0))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append([title])
    ws.cell(row=1, column=1).font = TITLE_FONT
    cols = ["#", "Naziv", "Kval /10", "Stil / Regija", "Tijelo 1-5",
            "Slatkoća 1-5", "Note", "Cijena €", "Dućan", "Serviranje", "Komentar"]
    ws.append(cols)
    style_header(ws, 2, len(cols))
    for i, d in enumerate(items, 1):
        ws.append([
            i,
            d["name"],
            d.get("qualityScore"),
            f"{d['style']} / {d.get('region','')}",
            d["body"],
            d["sweetness"],
            ", ".join(d["flavorTags"]),
            price_str(d.get("priceEUR")),
            d.get("shopHR", ""),
            (d.get("serving") or {}).get("best", ""),
            d["notes"]["hr"],
        ])
    widths = [4, 38, 9, 26, 10, 12, 30, 12, 18, 22, 60]
    for col, w in zip("ABCDEFGHIJK", widths):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
    wb.save(OUT / xlsx_name)
    print(f"{xlsx_name}: {len(items)} stavki")


def export_cigars():
    items = json.loads((DATA / "cigars.json").read_text(encoding="utf-8"))
    brands = json.loads((DATA / "brands.json").read_text(encoding="utf-8"))
    # sortiraj po snazi pa brendu — geografski stupci za filtriranje
    items.sort(key=lambda c: (c["strength"], c["brand"]))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cigare (geo dostupnost)"
    ws.append(["CIGARE — indeks po dostupnosti: filtriraj stupce HR/EU/USA/Svijet"])
    ws.cell(row=1, column=1).font = TITLE_FONT
    cols = ["Brand", "Linija", "Vitole (⏱ min)", "Zemlja", "Wrapper",
            "Snaga 1-5", "Tijelo 1-5", "Cijena €", "HR", "EU", "USA",
            "Svijet", "Note", "Komentar"]
    ws.append(cols)
    style_header(ws, 2, len(cols))
    for c in items:
        m = c["markets"]
        vitole = "; ".join(
            f"{v['name']} ({v['smokeTimeMin']}min{', ' + str(v['priceEUR']) + '€' if v.get('priceEUR') else ''})"
            for v in c.get("vitolas", []))
        ws.append([
            c["brand"], c["line"], vitole, c["country"],
            c["wrapper"], c["strength"], c["body"],
            price_str(c.get("priceEUR")),
            "✓" if "HR" in m else "", "✓" if "EU" in m else "",
            "✓" if "USA" in m else "", "✓" if "WW" in m else "",
            ", ".join(c["flavorTags"]), c["notes"]["hr"],
        ])
    widths = [16, 26, 55, 12, 26, 9, 9, 10, 5, 5, 6, 7, 28, 55]
    for col, w in zip(["A","B","C","D","E","F","G","H","I","J","K","L","M","N"], widths):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:N{ws.max_row}"

    # Sheet: Brand Index
    by_brand: dict[str, list] = {}
    for c in items:
        by_brand.setdefault(c["brand"], []).append(c)

    bcols = ["Brand", "Zemlja", "Founded", "# linija", "# vitola",
             "Additional Vitolas", "# search-only URL", "Najjeftinija €", "Blurb HR"]
    wb2 = wb.create_sheet("Brendovi")
    wb2.append(["CIGARE — indeks brendova (iz brands.json + cigars.json)"])
    wb2.cell(row=1, column=1).font = TITLE_FONT
    wb2.append(bcols)
    style_header(wb2, 2, len(bcols))
    for brand in sorted(by_brand.keys(), key=lambda s: s.casefold()):
        lines = by_brand[brand]
        info = brands.get(brand, {})
        named = [c for c in lines if c.get("line") != "Additional Vitolas"]
        vitola_n = sum(len(c.get("vitolas") or []) for c in lines)
        has_extra = any(c.get("line") == "Additional Vitolas" for c in lines)
        search_n = 0
        prices = []
        for c in lines:
            if c.get("priceEUR") is not None:
                prices.append(c["priceEUR"])
            for v in c.get("vitolas") or []:
                if v.get("priceEUR") is not None:
                    prices.append(v["priceEUR"])
                url = v.get("url") or ""
                if "?s=" in url or "post_type=product" in url:
                    search_n += 1
        blurb = (info.get("blurb") or {}).get("hr", "")
        if len(blurb) > 120:
            blurb = blurb[:117] + "..."
        wb2.append([
            brand,
            info.get("country", ""),
            info.get("founded", ""),
            len(named),
            vitola_n,
            "da" if has_extra else "",
            search_n,
            min(prices) if prices else "",
            blurb,
        ])
    bwidths = [22, 16, 12, 10, 10, 16, 16, 14, 60]
    for col, w in zip("ABCDEFGHI", bwidths):
        wb2.column_dimensions[col].width = w
    wb2.freeze_panes = "A3"
    wb2.auto_filter.ref = f"A2:I{wb2.max_row}"

    wb.save(OUT / "Cigare_Index.xlsx")
    print(f"Cigare_Index.xlsx: {len(items)} linija, {len(by_brand)} brendova")


if __name__ == "__main__":
    export_drinks("whiskies.json", "Whisky_Index.xlsx",
                  "WHISKY — rangirano po kvaliteti za sipping uz cigaru", "Whisky (rang)")
    export_drinks("brandies.json", "Konjak_Brandy_Index.xlsx",
                  "KONJAK / BRANDY — rangirano po kvaliteti za sipping uz cigaru", "Konjak-Brandy (rang)")
    export_drinks("coffees.json", "Kava_Index.xlsx",
                  "KAVA — profili za pairing s cigarom", "Kava (rang)")
    export_drinks("wines.json", "Vino_Index.xlsx",
                  "VINO - profili za pairing s cigarom", "Vino (rang)")
    export_cigars()
