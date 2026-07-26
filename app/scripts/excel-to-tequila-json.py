# -*- coding: utf-8 -*-
"""Ekstrakcija Tequila_Kolekcija_Checklist.xlsx -> tequilas.json.

Cita MASTER Ocjene, Serviranje + Cigare i Katalog allez+ecuga sheetove.
Zadrzava postojece unose iz tequilas.json koji nisu u MASTER-u (merge).

Pokretanje: python scripts/excel-to-tequila-json.py
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

from tequila_shared import (
    detect_age_tier,
    detect_category_type,
    detect_style_region,
    extract_abv,
    find_best_catalog_match,
    is_pairable,
    match_tokens,
    parse_price_eur,
    serving_for_style,
    slugify,
    token_overlap,
)

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT.parent / "Tequila_Kolekcija_Checklist.xlsx"
OUT = ROOT / "src" / "data" / "tequilas.json"
SEED = Path(__file__).resolve().parent / "seed" / "tequilas_seed.json"

SERVE_SCORE = {"++": 3, "+": 2, "~": 1, "x": 0}


def build_catalog_index(wb) -> list[dict]:
    index = []
    if "Katalog allez+ecuga" not in wb.sheetnames:
        return index
    for row in wb["Katalog allez+ecuga"].iter_rows(min_row=3, values_only=True):
        name, price, shop, url = row[0], row[1], row[2], row[3]
        if not name or not url:
            continue
        index.append({
            "name": str(name),
            "url": str(url),
            "shop": str(shop or ""),
            "price_eur": parse_price_eur(price),
            "tokens": match_tokens(str(name)),
        })
    return index


def build_serve_index(wb) -> list[dict]:
    rows = []
    if "Serviranje + Cigare" not in wb.sheetnames:
        return rows
    ws = wb["Serviranje + Cigare"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        name, neat, rocks, best, hint = (
            row[0], row[1], row[2], row[3], row[4] if len(row) > 4 else None
        )
        if not name or not best:
            continue
        rows.append({
            "tokens": match_tokens(str(name)),
            "serving": {
                "neat": SERVE_SCORE.get(str(neat).strip(), None) if neat else None,
                "rocks": SERVE_SCORE.get(str(rocks).strip(), None) if rocks else None,
                "best": str(best),
            },
            "cigarHintHr": str(hint) if hint else None,
        })
    return rows


def load_existing() -> list[dict]:
    if OUT.exists():
        return json.loads(OUT.read_text(encoding="utf-8-sig"))
    if SEED.exists():
        return json.loads(SEED.read_text(encoding="utf-8-sig"))
    return []


def find_existing(name: str, existing: list[dict]) -> dict | None:
    best, best_score = None, 0
    for item in existing:
        score = token_overlap(name, item["name"])
        if score > best_score:
            best, best_score = item, score
    return best if best and best_score >= 2 else None


def extract_tequilas(wb, existing: list[dict]) -> list[dict]:
    catalog = build_catalog_index(wb)
    serve_rows = build_serve_index(wb)
    ws = wb["MASTER Ocjene"]
    items = []
    seen_ids: set[str] = set()

    for row in ws.iter_rows(min_row=3, values_only=True):
        name, quality, region_col = row[0], row[1], row[2]
        if not name or quality is None:
            continue
        name_str = str(name)
        if name_str.startswith("VRH") or name_str.startswith("Odlican") or name_str.startswith("Value"):
            continue

        category_col = str(row[3] or "")
        age_tier = str(row[4] or detect_age_tier(name_str))
        body_col = row[5]
        sweet_col = row[6]
        price_raw = row[7]
        shop = str(row[8] or "")
        status = str(row[9]).strip() if row[9] and str(row[9]).strip() not in ("-", "") else None
        comment = str(row[10] or "") if len(row) > 10 else ""

        style, region, body, sweetness, tags = detect_style_region(name_str)
        category = category_col or detect_category_type(name_str)
        if region_col:
            region = str(region_col)
        if body_col is not None:
            body = int(body_col)
        if sweet_col is not None:
            sweetness = int(sweet_col)
        # Prefer age_tier as style when it is a known style key
        if age_tier in ("blanco", "reposado", "anejo", "extra-anejo", "mezcal"):
            style = age_tier

        abv = extract_abv(name_str) or 40.0
        quality_f = float(quality)

        my_tokens = match_tokens(name_str)
        best_match, best_score = None, 0
        for sr in serve_rows:
            overlap = len(my_tokens & sr["tokens"])
            if overlap > best_score:
                best_match, best_score = sr, overlap
        if best_match and best_score >= 2:
            serving = {k: v for k, v in best_match["serving"].items() if v is not None}
            if "best" not in serving:
                serving["best"] = serving_for_style(style)["best"]
        else:
            serving = serving_for_style(style)

        cat = find_best_catalog_match(name_str, catalog)
        price_eur = parse_price_eur(price_raw)
        price_url = cat["url"] if cat else None
        if price_eur is None and cat and cat.get("price_eur"):
            price_eur = cat["price_eur"]
        if cat and not shop:
            shop = cat.get("shop", "")

        prev = find_existing(name_str, existing)
        seed_notes = (prev or {}).get("notes") or {}
        hr_note = comment
        if (not hr_note or hr_note.startswith("Heuristika")) and seed_notes.get("hr"):
            hr_note = seed_notes["hr"]

        cid = (prev or {}).get("id") or ("tq-" + slugify(name_str))
        display_name = name_str
        if prev and token_overlap(name_str, prev["name"]) >= 3 and len(prev["name"]) < len(name_str):
            display_name = prev["name"]
            cid = prev["id"]
            if prev.get("abv"):
                abv = prev["abv"]
            if prev.get("flavorTags"):
                tags = prev["flavorTags"]
            if prev.get("style"):
                style = prev["style"]

        item = {
            "id": cid,
            "category": "tequila" if style != "mezcal" else "tequila",
            "name": display_name,
            "style": style,
            "region": region,
            "country": (prev or {}).get("country") or "Meksiko",
            "abv": abv,
            "body": body,
            "sweetness": sweetness,
            "flavorTags": tags,
            "qualityScore": quality_f,
            "priceEUR": {"min": price_eur, "max": price_eur} if price_eur else None,
            "priceApprox": price_eur is None,
            "shopHR": shop or "allez.hr",
            "priceUrl": price_url or (prev or {}).get("priceUrl"),
            "pairable": is_pairable(name_str, style, category, quality_f),
            "serving": serving,
            "cigarHint": (prev or {}).get("cigarHint"),
            "notes": {
                "hr": hr_note,
                "en": seed_notes.get("en", ""),
            },
        }
        if status:
            item["status"] = status

        if cid in seen_ids:
            continue
        seen_ids.add(cid)
        items.append(item)

    for prev in existing:
        if prev.get("id") in seen_ids:
            continue
        if any(token_overlap(prev["name"], it["name"]) >= 3 for it in items):
            continue
        items.append(prev)
        seen_ids.add(prev["id"])

    items.sort(key=lambda d: (-(d.get("qualityScore") or 0), d.get("name", "")))
    return items


def main() -> int:
    if not XLSX.exists():
        print(f"Missing {XLSX} — run build-tequila-excel.py first")
        return 1
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    existing = load_existing()
    tequilas = extract_tequilas(wb, existing)
    wb.close()
    OUT.write_text(json.dumps(tequilas, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"tequilas.json: {len(tequilas)} boca")
    with_url = sum(1 for t in tequilas if t.get("priceUrl"))
    print(f"  priceUrl: {with_url}/{len(tequilas)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
