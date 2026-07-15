# -*- coding: utf-8 -*-
"""Build Whisky_Kolekcija_Checklist.xlsx from whisky_catalog_raw.json + seed whiskies.json.

Pokretanje: python scripts/build-whisky-excel.py
"""
from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from serve_shared import load_corrections, resolve_serve_hint
from whisky_shared import (
    additive_status,
    catalog_index,
    cigar_hint_for_style,
    detect_coloring,
    detect_expression_type,
    detect_filter,
    detect_style_region,
    estimate_quality,
    extract_abv,
    find_best_catalog_match,
    format_price_eur,
    is_pairable,
    match_tokens,
    serving_for_style,
    token_overlap,
)

ROOT = Path(__file__).resolve().parent.parent
RAW = Path(__file__).resolve().parent / "output" / "whisky_catalog_raw.json"
SEED = ROOT / "src" / "data" / "whiskies.json"
XLSX = ROOT.parent / "Whisky_Kolekcija_Checklist.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="2C211A")
HEADER_FONT = Font(bold=True, color="C9A35C")
TITLE_FONT = Font(bold=True, size=12, color="9C4433")


def style_header(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center")


def load_seed() -> dict[str, dict]:
    if not SEED.exists():
        return {}
    items = json.loads(SEED.read_text(encoding="utf-8"))
    by_tokens: dict[str, dict] = {}
    for item in items:
        key = "|".join(sorted(match_tokens(item["name"])))
        by_tokens[key] = item
    return by_tokens


def find_seed(name: str, seeds: dict[str, dict]) -> dict | None:
    key = "|".join(sorted(match_tokens(name)))
    if key in seeds:
        return seeds[key]
    best, best_score = None, 0
    for seed in seeds.values():
        score = token_overlap(name, seed["name"])
        if score > best_score:
            best, best_score = seed, score
    return best if best and best_score >= 3 else None


def enrich_row(item: dict, seed: dict | None) -> dict:
    name = item["name"]
    style, region, body, sweetness, tags = detect_style_region(
        name, item.get("ecuga_category", "")
    )
    if seed:
        style = seed.get("style", style)
        region = normalize_region(str(seed.get("region", region)), style)
        body = seed.get("body", body)
        sweetness = seed.get("sweetness", sweetness)
        tags = seed.get("flavorTags", tags)
    expr = detect_expression_type(name)
    abv = seed.get("abv") if seed else extract_abv(name)
    coloring = detect_coloring(name, style, expr)
    filt = detect_filter(name)
    price = item.get("price_eur")
    if seed and seed.get("priceEUR"):
        p = seed["priceEUR"]
        if price is None:
            price = p.get("min")
    quality = estimate_quality(
        name, price, style, expr, abv,
        seed_score=seed.get("qualityScore") if seed else None,
    )
    note = (seed or {}).get("notes", {}).get("hr", "")
    if not note:
        note = f"Heuristika — {style}, {expr}"
    return {
        "name": name,
        "quality": quality,
        "style": style,
        "region": region,
        "body": body,
        "sweetness": sweetness,
        "tags": tags,
        "coloring": coloring,
        "filter": filt,
        "additive": additive_status(coloring, expr),
        "expr": expr,
        "abv": abv,
        "price": price,
        "shop": item.get("shop", ""),
        "url": item.get("url", ""),
        "note": note,
        "pairable": is_pairable(expr, style, quality),
        "seed": seed is not None,
    }


def build_svi_rang(catalog: list[dict], seeds: dict[str, dict]) -> list[dict]:
    rows = []
    for item in catalog:
        seed = find_seed(item["name"], seeds)
        rows.append(enrich_row(item, seed))
    rows.sort(key=lambda r: (-r["quality"], r["name"]))
    return rows


def select_master(all_rows: list[dict], seeds: dict[str, dict]) -> list[dict]:
    """Kurirani MASTER: svi seed + kvalitetni pairable do ~170."""
    seed_names = {s["name"] for s in seeds.values()}
    master: list[dict] = []
    seen: set[str] = set()

    for row in all_rows:
        if row["seed"] or any(token_overlap(row["name"], sn) >= 4 for sn in seed_names):
            key = row["name"][:60]
            if key not in seen:
                seen.add(key)
                master.append(row)

    for row in all_rows:
        if len(master) >= 170:
            break
        if row["quality"] >= 7.0 and row["pairable"]:
            key = row["name"][:60]
            if key not in seen:
                seen.add(key)
                master.append(row)

    master.sort(key=lambda r: (-r["quality"], r["name"]))
    return master


def group_by_type(rows: list[dict]) -> list[tuple[str, list[dict]]]:
    groups: dict[str, list[dict]] = defaultdict(list)
    style_labels = {
        "speyside-sherry": "Speyside sherry",
        "speyside-fruity": "Speyside fruity",
        "islay-peated": "Islay peated",
        "highland": "Highland",
        "campbeltown": "Campbeltown",
        "blended-scotch": "Blended Scotch",
        "bourbon": "Bourbon",
        "tennessee": "Tennessee",
        "rye": "Rye",
        "irish-pot-still": "Irish pot still",
        "irish-blend": "Irish blend",
        "japanese": "Japanese",
        "canadian": "Canadian",
        "world": "World whisky",
    }
    for row in rows:
        if not row["pairable"]:
            continue
        label = style_labels.get(row["style"], row["style"])
        groups[label].append(row)
    out = []
    for label in sorted(groups.keys()):
        items = sorted(groups[label], key=lambda r: -r["quality"])
        out.append((label, items))
    return out


def write_workbook(catalog: list[dict], seeds: dict[str, dict]) -> None:
    all_rows = build_svi_rang(catalog, seeds)
    master_rows = select_master(all_rows, seeds)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Katalog
    ws_k = wb.create_sheet("Katalog allez+ecuga")
    ws_k.append(["Katalog whisky — allez.hr + ecuga.com"])
    ws_k.cell(row=1, column=1).font = TITLE_FONT
    ws_k.append(["Naziv", "Cijena", "Web shop", "URL"])
    style_header(ws_k, 2, 4)
    for item in catalog:
        ws_k.append([
            item["name"],
            format_price_eur(item.get("price_eur")),
            item.get("shop", ""),
            item.get("url", ""),
        ])

    # Svi viskiji
    ws_s = wb.create_sheet("Svi viskiji (rang)")
    ws_s.append([
        "SVI VISKIJI RANGIRANI po kvaliteti (sipping uz cigaru) | allez.hr + ecuga.com",
    ])
    ws_s.append(["#", "Whisky", "Kval /10", "Tip / Regija", "Bilješka", "Cijena", "Shop"])
    style_header(ws_s, 2, 7)
    for i, row in enumerate(all_rows, 1):
        ws_s.append([
            i,
            row["name"],
            row["quality"],
            f"{row['region']}",
            row["note"][:80],
            format_price_eur(row["price"]),
            row["shop"],
        ])

    # MASTER
    ws_m = wb.create_sheet("MASTER Ocjene")
    ws_m.append(["MASTER — REKALIBRIRANO za SIPPING UZ CIGARU. Rangirano po kvaliteti."])
    ws_m.append([
        "Whisky", "Kval /10", "Tip / Regija", "Boja (E150)",
        "Filter", "Expression", "Cijena €", "Dućan", "Status", "Komentar",
    ])
    style_header(ws_m, 2, 10)
    current_tier = None
    for row in master_rows:
        tier = (
            "VRH za cigaru (8.0-10)" if row["quality"] >= 8
            else "Odličan sipper (7-8)" if row["quality"] >= 7
            else "Value / solidan (6-7)"
        )
        if tier != current_tier:
            ws_m.append([tier, None, None, None, None, None, None, None, None, None])
            current_tier = tier
        ws_m.append([
            row["name"],
            row["quality"],
            f"{row['region']}",
            row["coloring"],
            row["filter"],
            row["expr"],
            format_price_eur(row["price"]),
            row["shop"],
            "META" if row["seed"] else "",
            row["note"],
        ])

    # Po tipu
    ws_t = wb.create_sheet("Po tipu (kupnja)")
    ws_t.append([
        "PO TIPU — vodič za kupnju (1-2 iz grupe). Rangirano po kvaliteti unutar grupe.",
    ])
    ws_t.append([
        "Whisky", "Kval /10", "Expression", "Cijena €", "Dućan/izvor", "Status / bilješka",
    ])
    style_header(ws_t, 2, 6)
    for label, items in group_by_type(master_rows):
        ws_t.append([f"{label}  ({len(items)})", None, None, None, None, None])
        for row in items[:25]:
            ws_t.append([
                row["name"],
                row["quality"],
                row["expr"],
                format_price_eur(row["price"]),
                row["shop"],
                row["note"][:100],
            ])

    # Serviranje + Cigare
    ws_sv = wb.create_sheet("Serviranje + Cigare")
    ws_sv.append(["SERVIRANJE + CIGARE — profili stilova za whisky"])
    ws_sv.append([
        "Profil / primjer", "Neat", "Voda", "Rocks", "Best", "Cigar hint",
    ])
    style_header(ws_sv, 2, 6)
    profiles = [
        ("Speyside sherry (Aberlour, GlenDronach)", "speyside-sherry"),
        ("Islay peated (Ardbeg, Laphroaig)", "islay-peated"),
        ("Bourbon (Buffalo Trace, Four Roses)", "bourbon"),
        ("Irish pot still (Redbreast, Powers)", "irish-pot-still"),
        ("Japanese (Nikka, Yamazaki)", "japanese"),
        ("Blended Scotch (JW Black, Chivas)", "blended-scotch"),
    ]
    serve_map = {"3": "++", "2": "+", "1": "~", "0": "x"}
    for label, style in profiles:
        s = serving_for_style(style, 46, "single-malt")
        ws_sv.append([
            label,
            serve_map.get(str(s["neat"]), "+"),
            serve_map.get(str(s["water"]), "+"),
            serve_map.get(str(s["rocks"]), "~"),
            s["best"],
            resolve_serve_hint(label, style, cigar_hint_for_style),
        ])
    for row in master_rows[:40]:
        s = serving_for_style(row["style"], row.get("abv"), row["expr"])
        ws_sv.append([
            row["name"],
            serve_map.get(str(s["neat"]), "+"),
            serve_map.get(str(s["water"]), "+"),
            serve_map.get(str(s["rocks"]), "~"),
            s["best"],
            resolve_serve_hint(row["name"], row["style"], cigar_hint_for_style),
        ])

    # Kolekcija plan
    ws_p = wb.create_sheet("Kolekcija (plan)")
    ws_p.append(["TIER plan — whisky kolekcija (po stilu, bez duplikata)"])
    ws_p.append([
        "✓", "Tier", "Stil meta", "Boca meta", "Profil", "Izvor cijene", "Moja ocjena", "Bilješke",
    ])
    style_header(ws_p, 2, 8)
    tiers = [
        ("TIER 1 — must have", "Islay peated", "Ardbeg 10 / Laphroaig 10", "Dim + medicinski", "allez.hr", "8+"),
        ("TIER 1 — must have", "Speyside sherry", "GlenDronach 12 / Aberlour 12", "Sherry + kakao", "allez.hr", "8+"),
        ("TIER 2 — core", "Bourbon", "Buffalo Trace / Four Roses SiB", "Karamela + vanilija", "allez.hr", "7.5+"),
        ("TIER 2 — core", "Irish pot still", "Redbreast 12 / Powers John's Lane", "Kremasto + začini", "allez.hr", "8+"),
        ("TIER 3 — explore", "Japanese", "Nikka Yoichi / Hakushu", "Elegantno + cvjetno", "allez.hr", "7.5+"),
        ("TIER 3 — explore", "Campbeltown", "Springbank 10", "Složeno + sol", "allez.hr", "8.5+"),
        ("TIER 4 — luxury", "Sherry CS", "Aberlour A'bunadh / Glenfarclas 105", "Cask strength", "allez.hr", "8+"),
    ]
    for t in tiers:
        ws_p.append(["", *t, "", ""])

    # Vodič
    ws_g = wb.create_sheet("Vodic (sazetak)")
    guide = [
        "WHISKY VODIČ — sažetak za sipping uz cigaru",
        "",
        "E150a bojenje: blended Scotch često koristi karamel (E150). Single malt obično natural.",
        "Chill-filter: NCF oznaka = nije chill-filtered (više tijela, bolje uz jaču cigaru).",
        "NAS vs age: age statement pouzdaniji za očekivani profil; NAS ne znači loše, ali provjeri.",
        "Flavoured / RTD / cocktail viski: u punom katalogu, NE u MASTER/app (pairable=false).",
        "Pairing: maduro wrapper → sherry/bourbon; Habano/corojo → Islay/rye; Connecticut → Irish/Japanese.",
        "",
        "Izvori cijena: allez.hr (primarni) + ecuga.com. Regeneriraj: scrape-whisky-catalog.py",
    ]
    for line in guide:
        ws_g.append([line])

    wb.save(XLSX)
    print(f"Wrote {XLSX}")
    print(f"  Katalog: {len(catalog)} | Svi: {len(all_rows)} | MASTER: {len(master_rows)}")


def main() -> int:
    if not RAW.exists():
        print(f"Missing {RAW} — run scrape-whisky-catalog.py first")
        return 1
    catalog = json.loads(RAW.read_text(encoding="utf-8"))
    seeds_list = json.loads(SEED.read_text(encoding="utf-8")) if SEED.exists() else []
    seeds = {s["name"]: s for s in seeds_list}
    write_workbook(catalog, seeds)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
